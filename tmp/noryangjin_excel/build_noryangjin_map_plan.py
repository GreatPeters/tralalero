from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT / "outputs" / "noryangjin-map-plan"
OUTPUT_PATH = OUTPUT_DIR / "noryangjin_maptool1_analysis_and_expansion_plan.xlsx"
CONCEPT_PATH = OUTPUT_DIR / "noryangjin_expansion_concept.png"
QA_DIR = OUTPUT_DIR / "qa"
SOURCE_SCENE = "Assets/ShooterSurvival/Scenes/Tools/Noryangjin_MapTool_Mode.unity"
SOURCE_SHA256 = "AB30AE1AC3B63EA322839E246E342BF800C7C0EB37AD90A690DE68A93DA76795"


COLORS = {
    "navy": "132A3A",
    "navy2": "1E3A4C",
    "cyan": "35C2D8",
    "water": "0E6574",
    "water_dark": "084B58",
    "existing": "6B4423",
    "existing_edge": "A06A3A",
    "proposed": "D97706",
    "proposed_light": "F59E0B",
    "clear": "FDE68A",
    "building": "334155",
    "object": "2E8B57",
    "gameplay": "8B5CF6",
    "danger": "C2413A",
    "scenery": "2563A6",
    "white": "FFFFFF",
    "ink": "17202A",
    "muted": "6B7280",
    "line": "CBD5E1",
    "panel": "E8F2F4",
    "panel2": "F4F8F9",
    "pale_orange": "FFF1D6",
    "pale_blue": "E0F2FE",
    "pale_green": "DCFCE7",
    "pale_red": "FEE2E2",
}


thin = Side(style="thin", color=COLORS["line"])
medium_navy = Side(style="medium", color=COLORS["navy"])


ROAD_ROWS = [
    ("Road_Basic", -7.675, -10.300, 0, "서측 세로축/상단 접점"),
    ("Road_Basic", -7.675, -21.550, 0, "서측 세로축"),
    ("Road_Basic", -7.675, -32.800, 0, "서측 세로축"),
    ("Road_Basic", -7.675, -44.050, 0, "서측 세로축"),
    ("Road_Basic", -7.675, -55.525, 0, "서측 세로축"),
    ("Road_Basic", -7.675, -66.775, 0, "서측 세로축"),
    ("Road_Basic", -7.450, -78.025, 0, "서측 세로축"),
    ("Road_Basic", -7.450, -89.275, 0, "서측 세로축"),
    ("Road_Basic", -7.450, -100.525, 0, "서측 세로축"),
    ("Road_RightTurn", -2.850, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 8.400, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 19.650, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 30.900, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 42.150, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 53.400, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 64.650, -10.450, 90, "북측 횡축"),
    ("Road_RightTurn", 75.900, -10.450, 90, "북측 횡축"),
    ("Road_LeftTurn", -14.250, -111.775, 270, "남측 짧은 횡축"),
    ("Road_LeftTurn", -3.000, -111.775, 270, "남측 짧은 횡축"),
    ("Road_LeftTurn", 8.275, -111.750, 270, "남측 짧은 횡축"),
    ("Road_LeftTurn", 19.525, -111.750, 270, "남측 짧은 횡축"),
]


OBJECT_GROUPS = [
    ("해수 배경", 98, -149.475, 77.775, -141.825, 16.350, "배경"),
    ("생선 부산물 바닥 장식", 13, -38.230, 68.645, -123.800, 5.575, "장식"),
    ("갈매기 횃대 포스트", 12, -14.222, 30.328, -111.415, -3.640, "장식"),
    ("우천 시장 바닥 배경", 11, -21.250, 76.625, -60.825, -14.025, "배경"),
    ("복어 적 프리뷰", 8, -31.510, 32.840, -115.650, 0.000, "게임플레이"),
    ("문어 수조", 7, -16.375, -16.375, -32.650, -12.400, "주요 오브젝트"),
    ("게 수조", 7, -5.935, -5.935, -32.650, -12.400, "주요 오브젝트"),
    ("구명부환", 3, -16.550, -5.300, -86.300, -1.475, "소품"),
    ("해산물 진열대", 2, -17.725, -4.000, -65.725, -53.825, "건물"),
    ("회 식당 전면", 2, -17.780, -4.030, -43.265, -42.590, "건물"),
    ("어시장 상점 파사드", 2, -17.825, -3.800, -64.675, -55.075, "건물"),
    ("닻 소품", 1, -12.470, -12.470, -83.360, -83.360, "소품"),
    ("어선", 1, -24.200, -24.200, -77.940, -77.940, "배경 오브젝트"),
    ("항만 신호 갠트리", 1, -10.735, -10.735, -93.000, -93.000, "게임플레이"),
    ("원경 마을", 1, -97.650, -97.650, -62.300, -62.300, "배경 오브젝트"),
    ("해상 부표", 1, -21.825, -21.825, -124.650, -124.650, "배경 오브젝트"),
]


PLACEMENTS = [
    ("P1", "남측", "길", "하단 부두 동쪽 연장", "X3~X7 / Z9", 5, "막힌 끝을 전투 광장으로 연결", "최소 3칸 폭 유지", "미정", "기존 남측 4개 모듈 뒤에 연결"),
    ("P1", "동측", "길", "동측 세로 서비스 부두", "X7 / Z1~Z8", 8, "북측과 남측을 잇는 순환 동선", "교차점 4칸 폭", "미정", "상단 끝과 남측 광장에 접속"),
    ("P1", "중앙", "길", "중간 연결로", "X1~X6 / Z5", 6, "두 개의 짧은 순환 선택지 생성", "중앙 3칸 비움", "미정", "서측 밀집 구간의 우회로"),
    ("P1", "남측", "길", "전투 광장 확장 데크", "X3~X7 / Z8~Z11", 12, "웨이브·보스·보상 이벤트 공간", "중심 3×2칸 비움", "미정", "가장자리만 소품 배치"),
    ("P1", "전체", "규칙", "주행로 폭 규칙", "모든 길", 0, "이동과 사격 가독성 확보", "직선 3칸 / 교차점 4칸", "미정", "셀 색상 노랑 구역은 항상 비움"),
    ("P1", "북측", "오브젝트", "입구 신호 갠트리", "X7 / Z-1", 1, "시작 방향과 구역 전환 명확화", "길 정중앙 높이 여유 확보", "미정", "현재 갠트리와 역할 중복 주의"),
    ("P2", "북측", "오브젝트", "상자·그물·구명부환 군집", "X1~X6 / Z-1", 8, "긴 횡축의 비어 보이는 구간 보강", "2칸마다 한쪽 어깨만", "미정", "좌우 번갈아 배치"),
    ("P1", "서측", "조정", "수조 반복 군집화", "X-1~X1 / Z1~Z4", 4, "14개 반복 수조를 4개 장면 군집으로 읽히게 함", "각 군집 사이 2칸 비움", "미정", "삭제안이 아니라 재배치 가이드"),
    ("P2", "서측", "오브젝트", "시장 가판·천막 변주", "X-1~X1 / Z3~Z6", 5, "서측 시장 정체성 강화", "시야 높이 1.5칸 이하", "미정", "파란/주황 천막 교차"),
    ("P1", "중앙", "오브젝트", "경매장 가장자리 마커", "X2~X5 / Z4~Z6", 6, "전투 공간의 경계만 암시", "중앙 연결로 침범 금지", "미정", "바닥선·낮은 상자 위주"),
    ("P2", "중앙", "게임플레이", "코인·보상 라인", "X2~X5 / Z5", 1, "플레이어를 우회로로 유도", "직선 가시성 유지", "미정", "웨이브 이후 활성화 권장"),
    ("P2", "동측", "오브젝트", "계류주·램프·갈매기 포스트", "X7~X8 / Z2~Z7", 10, "긴 동측 부두의 리듬 형성", "한 모듈당 최대 1개", "미정", "현재 포스트 반복보다 간격 확대"),
    ("P2", "동측", "배경", "소형 어선", "X8 / Z3, Z7", 2, "항만 공간감과 방향 기준점", "충돌 영역은 물 위", "미정", "플레이 가능 데크 밖"),
    ("P2", "남측", "오브젝트", "생선 카트·지게차·배럴", "광장 외곽", 7, "전투 광장 실루엣 강화", "중심 3×2칸 비움", "미정", "큰 오브젝트는 모서리에만"),
    ("P1", "남측", "게임플레이", "최종 게이트", "X5 / Z12", 1, "보스·스테이지 종료 지점 명확화", "게이트 전 4칸 비움", "미정", "세로축 중심 정렬"),
    ("P2", "동측", "길", "보상용 짧은 사이드 도크", "X8~X10 / Z5", 3, "선택 보상 알코브", "복귀 동선 2칸 확보", "미정", "막다른 길 길이 제한"),
    ("P3", "전체", "장식", "부산물·부표 희소 배치", "어깨/수면", 12, "빈 공간의 질감 보강", "연속 2개 배치 금지", "미정", "현재 13개 부산물은 분산 유지"),
    ("P1", "전체", "검증", "통로 막힘 점검", "모든 접점", 0, "오브젝트 추가 후 플레이 가능성 확인", "렌더러 경계와 3칸 통로 대조", "미정", "Unity 반영 전에 엑셀로 1차 검토"),
]


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def apply_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = fill(COLORS["navy"])
    cell.font = Font(name="맑은 고딕", size=20, bold=True, color=COLORS["white"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_col)
    sub = ws.cell(3, 1, subtitle)
    sub.fill = fill(COLORS["navy2"])
    sub.font = Font(name="맑은 고딕", size=10, color="D7EEF2")
    sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 28


def style_header(row_cells: Iterable, color: str = COLORS["navy2"]) -> None:
    for cell in row_cells:
        cell.fill = fill(color)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium_navy)


def add_table(ws, ref: str, name: str, style_name: str = "TableStyleMedium2") -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_sheet_defaults(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "00_읽는법"
    ws.sheet_properties.tabColor = COLORS["cyan"]
    set_sheet_defaults(ws)
    apply_title(
        ws,
        "노량진 맵툴1 분석 · 추가 동선 설계안",
        "Unity 원본 씬은 수정하지 않고, 현재 배치와 추가 길·오브젝트 방향을 엑셀 셀 도면으로 정리한 검토용 문서입니다.",
        16,
    )

    cards = [
        (1, 4, "현재 도로", "='05_원본근거'!J4", "개 모듈"),
        (5, 8, "현재 오브젝트", "='05_원본근거'!J5", "개"),
        (9, 12, "실질 비배경", "='05_원본근거'!J8", "개"),
        (13, 16, "제안 핵심", "2", "개 순환 루프"),
    ]
    for start_col, end_col, label, formula, unit in cards:
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        ws.merge_cells(start_row=6, start_column=start_col, end_row=7, end_column=end_col)
        ws.merge_cells(start_row=8, start_column=start_col, end_row=8, end_column=end_col)
        ws.cell(5, start_col, label)
        ws.cell(6, start_col, formula)
        ws.cell(8, start_col, unit)
        for row in range(5, 9):
            c = ws.cell(row, start_col)
            c.fill = fill(COLORS["panel"] if start_col < 13 else COLORS["pale_orange"])
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(5, start_col).font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["navy"])
        ws.cell(6, start_col).font = Font(name="맑은 고딕", size=24, bold=True, color=COLORS["navy"])
        ws.cell(8, start_col).font = Font(name="맑은 고딕", size=9, color=COLORS["muted"])
        ws.cell(5, start_col).border = Border(top=medium_navy, left=medium_navy, right=medium_navy)
        ws.cell(8, start_col).border = Border(bottom=medium_navy, left=medium_navy, right=medium_navy)

    legends = [
        ("현재 길", COLORS["existing"]),
        ("추가 길", COLORS["proposed"]),
        ("항상 비움", COLORS["clear"]),
        ("건물/시장", COLORS["building"]),
        ("소품", COLORS["object"]),
        ("게임플레이", COLORS["gameplay"]),
        ("적/위험", COLORS["danger"]),
        ("배경/선박", COLORS["scenery"]),
    ]
    ws.cell(10, 1, "범례")
    ws.cell(10, 1).font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["navy"])
    for idx, (label, color) in enumerate(legends, start=2):
        ws.cell(10, idx, label)
        ws.cell(10, idx).fill = fill(color)
        ws.cell(10, idx).font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["ink"] if color == COLORS["clear"] else COLORS["white"])
        ws.cell(10, idx).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A11:P11")
    ws["A11"] = "권장 읽기 순서: 01_현재맵 → 02_추가제안 → 03_배치목록 → 04_분석요약 → 05_원본근거"
    ws["A11"].fill = fill(COLORS["panel2"])
    ws["A11"].font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["navy"])
    ws["A11"].alignment = Alignment(horizontal="center")

    image = XLImage(str(CONCEPT_PATH))
    image.width = 1120
    image.height = 630
    ws.add_image(image, "A13")
    for col in range(1, 17):
        ws.column_dimensions[chr(64 + col)].width = 10.5
    for row in range(13, 43):
        ws.row_dimensions[row].height = 16
    ws.freeze_panes = "A4"
    ws.print_area = "A1:P43"
    ws.page_setup.orientation = "landscape"


def map_cell(ws, x: int, z: int):
    # Map area: B:Q, rows 7:21. X=-3..12, Z=-2..12.
    return ws.cell(row=7 + (z + 2), column=2 + (x + 3))


def setup_map_sheet(ws, title: str, subtitle: str, proposed: bool) -> None:
    set_sheet_defaults(ws)
    ws.sheet_properties.tabColor = COLORS["proposed"] if proposed else COLORS["existing"]
    apply_title(ws, title, subtitle, 23)
    ws.freeze_panes = "B7"
    ws.sheet_view.zoomScale = 85

    legend = [
        ("물", COLORS["water"]),
        ("현재 길", COLORS["existing"]),
        ("추가 길", COLORS["proposed"]),
        ("비움", COLORS["clear"]),
        ("건물", COLORS["building"]),
        ("소품", COLORS["object"]),
        ("게임", COLORS["gameplay"]),
        ("위험", COLORS["danger"]),
    ]
    ws.cell(4, 1, "범례")
    ws.cell(4, 1).font = Font(name="맑은 고딕", bold=True, color=COLORS["navy"])
    for i, (label, color) in enumerate(legend, start=2):
        ws.cell(4, i, label)
        ws.cell(4, i).fill = fill(color)
        ws.cell(4, i).font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["ink"] if color == COLORS["clear"] else COLORS["white"])
        ws.cell(4, i).alignment = Alignment(horizontal="center")

    ws.cell(6, 1, "Z↓ / X→")
    ws.cell(6, 1).fill = fill(COLORS["navy2"])
    ws.cell(6, 1).font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["white"])
    for x in range(-3, 13):
        c = map_cell(ws, x, -2)
        header = ws.cell(6, c.column, x)
        header.fill = fill(COLORS["navy2"])
        header.font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["white"])
        header.alignment = Alignment(horizontal="center")

    for z in range(-2, 13):
        row = 7 + (z + 2)
        label = ws.cell(row, 1, z)
        label.fill = fill(COLORS["navy2"])
        label.font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["white"])
        label.alignment = Alignment(horizontal="center", vertical="center")
        for x in range(-3, 13):
            c = map_cell(ws, x, z)
            c.fill = fill(COLORS["water"] if (x + z) % 2 == 0 else COLORS["water_dark"])
            c.font = Font(name="맑은 고딕", size=8, bold=True, color=COLORS["white"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(2, 18):
        ws.column_dimensions[ws.cell(6, col).column_letter].width = 6.0
    ws.column_dimensions["A"].width = 8
    for row in range(7, 22):
        ws.row_dimensions[row].height = 31

    # Current route: north east-west pier, west spine, short lower pier.
    current_roads = {(x, 0) for x in range(0, 8)} | {(0, z) for z in range(0, 9)} | {(x, 9) for x in range(-1, 3)}
    for x, z in current_roads:
        c = map_cell(ws, x, z)
        c.fill = fill(COLORS["existing"])
        c.border = Border(left=Side(style="medium", color=COLORS["existing_edge"]), right=Side(style="medium", color=COLORS["existing_edge"]), top=Side(style="medium", color=COLORS["existing_edge"]), bottom=Side(style="medium", color=COLORS["existing_edge"]))
        c.value = "길"
        c.comment = Comment("현재 Noryangjin_MapTool_Mode.unity 도로 골격", "User")

    ws.merge_cells("S6:W6")
    ws["S6"] = "설계 메모"
    ws["S6"].fill = fill(COLORS["navy2"])
    ws["S6"].font = Font(name="맑은 고딕", bold=True, color=COLORS["white"])
    ws["S6"].alignment = Alignment(horizontal="center")
    for col in range(19, 24):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.print_area = "A1:W23"
    ws.page_setup.orientation = "landscape"


def place_marker(ws, x: int, z: int, label: str, color: str, comment: str) -> None:
    c = map_cell(ws, x, z)
    c.value = label
    c.fill = fill(color)
    c.font = Font(name="맑은 고딕", size=8, bold=True, color=COLORS["white"])
    c.comment = Comment(comment, "User")


def build_current_map(wb: Workbook) -> None:
    ws = wb.create_sheet("01_현재맵")
    setup_map_sheet(
        ws,
        "현재 맵 · 좌표 기반 셀 도면",
        "한 셀은 도로 모듈 약 1개(약 11.25 Unity units)를 뜻합니다. 실제 배치는 21개 도로와 170개 Props 직접 자식으로 구성됩니다.",
        proposed=False,
    )
    markers = [
        (-1, 0, "수조", COLORS["building"], "문어/게 수조 반복 구간"),
        (1, 1, "수조", COLORS["building"], "문어/게 수조 반복 구간"),
        (-1, 2, "시장", COLORS["building"], "어시장 상점 파사드"),
        (1, 2, "식당", COLORS["building"], "회 식당 전면"),
        (-1, 4, "진열", COLORS["building"], "해산물 진열대"),
        (1, 4, "수조", COLORS["building"], "수조 반복 구간"),
        (-2, 6, "어선", COLORS["scenery"], "서측 수면의 항구 어선"),
        (-1, 6, "닻", COLORS["object"], "닻 소품"),
        (1, 7, "갠트리", COLORS["gameplay"], "항만 신호 갠트리"),
        (-1, 9, "부표", COLORS["scenery"], "해상 부표/구명부환"),
        (2, -1, "포스트", COLORS["object"], "갈매기 횃대 포스트"),
        (4, -1, "포스트", COLORS["object"], "갈매기 횃대 포스트"),
        (6, -1, "적", COLORS["danger"], "복어 적 프리뷰"),
        (3, 1, "적", COLORS["danger"], "복어 적 프리뷰"),
        (3, 10, "스크랩", COLORS["object"], "생선 부산물 장식"),
    ]
    for marker in markers:
        place_marker(ws, *marker)

    notes = [
        "핵심 형태: 북측 긴 횡축 → 서측 세로축 → 남측 짧은 끝.",
        "170개 중 109개가 해수/우천 바닥 배경이라 체감 오브젝트는 61개.",
        "시장·수조는 서측 Z1~Z6에 집중되어 북측과 남측이 상대적으로 비어 보임.",
        "수조 14개, 부산물 13개, 포스트 12개가 반복되어 다양성이 낮게 읽힘.",
        "남측은 4개 모듈에서 끝나 이벤트·보스 공간으로 확장 여지가 큼.",
    ]
    for note_idx, note in enumerate(notes):
        idx = 7 + note_idx * 2
        ws.merge_cells(start_row=idx, start_column=19, end_row=idx + 1, end_column=23)
        c = ws.cell(idx, 19, f"• {note}")
        c.fill = fill(COLORS["panel2"])
        c.font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])
        c.alignment = Alignment(wrap_text=True, vertical="center")


def build_proposal_map(wb: Workbook) -> None:
    ws = wb.create_sheet("02_추가제안")
    setup_map_sheet(
        ws,
        "추가 제안 · 2개 루프 + 남측 전투 광장",
        "갈색은 현재 길, 주황은 추가 길, 노랑은 오브젝트를 두지 않는 전투·회전 공간입니다. 이 시트는 제안 도면이며 Unity 씬에는 반영하지 않았습니다.",
        proposed=True,
    )

    proposed_cells = (
        {(x, 9) for x in range(3, 8)}
        | {(7, z) for z in range(1, 9)}
        | {(x, 5) for x in range(1, 7)}
        | {(x, z) for x in range(3, 8) for z in range(8, 12)}
        | {(x, 5) for x in range(8, 11)}
    )
    for x, z in proposed_cells:
        c = map_cell(ws, x, z)
        c.fill = fill(COLORS["proposed"] if (x + z) % 2 == 0 else COLORS["proposed_light"])
        c.font = Font(name="맑은 고딕", size=8, bold=True, color=COLORS["white"])
        c.value = "추가"
        c.comment = Comment("추가 제안 길/데크. Unity 원본 미반영.", "User")

    for x, z in {(4, 9), (5, 9), (6, 9), (4, 10), (5, 10), (6, 10), (3, 5), (4, 5), (5, 5)}:
        c = map_cell(ws, x, z)
        c.fill = fill(COLORS["clear"])
        c.font = Font(name="맑은 고딕", size=8, bold=True, color=COLORS["ink"])
        c.value = "비움"
        c.comment = Comment("전투/회전/시야 확보를 위해 항상 비우는 공간", "User")

    markers = [
        (7, -1, "입구문", COLORS["gameplay"], "북측 입구 신호 갠트리"),
        (5, -1, "상자", COLORS["object"], "상자·그물·구명부환 군집"),
        (2, -1, "그물", COLORS["object"], "북측 시장 어깨 장식"),
        (-1, 1, "수조군", COLORS["building"], "수조를 장면 군집으로 정리"),
        (1, 3, "천막", COLORS["building"], "시장 가판·천막 변주"),
        (2, 4, "경매", COLORS["gameplay"], "경매장 가장자리 마커"),
        (6, 4, "경매", COLORS["gameplay"], "경매장 가장자리 마커"),
        (8, 3, "어선", COLORS["scenery"], "동측 소형 어선"),
        (8, 6, "계류", COLORS["object"], "계류주·램프·포스트"),
        (9, 7, "부표", COLORS["scenery"], "동측 수면 부표"),
        (3, 11, "카트", COLORS["object"], "남측 광장 가장자리 카트"),
        (7, 11, "지게차", COLORS["object"], "남측 광장 모서리 지게차"),
        (5, 12, "최종문", COLORS["gameplay"], "보스/스테이지 종료 게이트"),
        (10, 4, "보상", COLORS["gameplay"], "선택 보상 사이드 도크"),
    ]
    for marker in markers:
        place_marker(ws, *marker)

    notes = [
        "1순위: 남측 짧은 끝을 동쪽으로 연장해 넓은 전투 광장 조성.",
        "1순위: 동측 세로 부두와 중앙 연결로로 두 개의 순환 선택지 생성.",
        "서측은 오브젝트를 더 쌓기보다 수조·가판을 4개 장면 군집으로 정리.",
        "북측은 입구/시장, 중앙은 경매/전투, 동측은 선착장, 남측은 보스 구역.",
        "길 중앙은 항상 3칸, 교차점·게이트 앞은 4칸 이상 비워 둠.",
    ]
    for note_idx, note in enumerate(notes):
        idx = 7 + note_idx * 2
        ws.merge_cells(start_row=idx, start_column=19, end_row=idx + 1, end_column=23)
        c = ws.cell(idx, 19, f"{note_idx + 1}. {note}")
        c.fill = fill(COLORS["pale_orange"] if idx <= 8 else COLORS["panel2"])
        c.font = Font(name="맑은 고딕", size=9, bold=idx <= 8, color=COLORS["ink"])
        c.alignment = Alignment(wrap_text=True, vertical="center")


def build_placements(wb: Workbook) -> None:
    ws = wb.create_sheet("03_배치목록")
    ws.sheet_properties.tabColor = COLORS["object"]
    set_sheet_defaults(ws)
    apply_title(
        ws,
        "추가 길 · 오브젝트 실행 목록",
        "수량과 상태는 수정 가능하며, 04_분석요약의 제안 지표가 이 표를 참조합니다.",
        10,
    )
    headers = ["우선순위", "구역", "유형", "추가/조정 항목", "권장 위치", "수량", "게임플레이 목적", "비우기 규칙", "상태", "담당/비고"]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    style_header(ws[5])
    for row_idx, values in enumerate(PLACEMENTS, start=6):
        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row_idx, col_idx, value)
            c.font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 38
        if values[0] == "P1":
            ws.cell(row_idx, 1).fill = fill(COLORS["pale_red"])
            ws.cell(row_idx, 1).font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["danger"])
        elif values[0] == "P2":
            ws.cell(row_idx, 1).fill = fill(COLORS["pale_orange"])
        else:
            ws.cell(row_idx, 1).fill = fill(COLORS["panel2"])
    end_row = 5 + len(PLACEMENTS)
    add_table(ws, f"A5:J{end_row}", "NoryangjinPlacementPlan", "TableStyleMedium4")

    priority_validation = DataValidation(type="list", formula1='"P1,P2,P3"', allow_blank=False)
    status_validation = DataValidation(type="list", formula1='"미정,검토중,확정,보류"', allow_blank=False)
    ws.add_data_validation(priority_validation)
    ws.add_data_validation(status_validation)
    priority_validation.add(f"A6:A{end_row}")
    status_validation.add(f"I6:I{end_row}")

    widths = [10, 10, 12, 26, 18, 8, 34, 28, 10, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(5, idx).column_letter].width = width
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:J{end_row}"
    ws.print_area = f"A1:J{end_row}"
    ws.page_setup.orientation = "landscape"


def build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("04_분석요약")
    ws.sheet_properties.tabColor = COLORS["gameplay"]
    set_sheet_defaults(ws)
    apply_title(
        ws,
        "현재 밀도 분석 · 추가 설계 판단",
        "원본 근거 시트와 배치 목록을 참조하는 수식 기반 요약입니다.",
        12,
    )

    cards = [
        (1, 3, "현재 도로", "='05_원본근거'!J4"),
        (4, 6, "현재 Props", "='05_원본근거'!J5"),
        (7, 9, "배경 Props", "=SUM('05_원본근거'!J6:J7)"),
        (10, 12, "실질 비배경", "='05_원본근거'!J8"),
    ]
    for start, end, label, formula in cards:
        ws.merge_cells(start_row=5, start_column=start, end_row=5, end_column=end)
        ws.merge_cells(start_row=6, start_column=start, end_row=7, end_column=end)
        ws.cell(5, start, label)
        ws.cell(6, start, formula)
        for row in range(5, 8):
            c = ws.cell(row, start)
            c.fill = fill(COLORS["panel"])
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(5, start).font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["navy"])
        ws.cell(6, start).font = Font(name="맑은 고딕", size=22, bold=True, color=COLORS["navy"])

    ws["A9"] = "현재 구성"
    ws["A9"].font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["navy"])
    headers = ["구분", "개수", "전체 비중", "해석"]
    for col, header in enumerate(headers, start=1):
        ws.cell(10, col, header)
    style_header(ws[10][:4])
    composition = [
        ("해수 배경", "='05_원본근거'!J6", "맵 면적 확보용 반복 배경"),
        ("우천 바닥", "='05_원본근거'!J7", "길 주변 바닥 반복 배경"),
        ("수조", "=SUM('05_원본근거'!B10:B11)", "서측 초반에 집중된 반복 오브젝트"),
        ("부산물", "='05_원본근거'!B6", "희소 장식이지만 수량은 큼"),
        ("포스트", "='05_원본근거'!B7", "간격 조정이 필요한 반복 장식"),
        ("적 프리뷰", "='05_원본근거'!B9", "경로 전반에 분산"),
        ("건물 3종", "=SUM('05_원본근거'!B13:B15)", "정체성 오브젝트 수가 적음"),
    ]
    for row_idx, (name, formula, note) in enumerate(composition, start=11):
        ws.cell(row_idx, 1, name)
        ws.cell(row_idx, 2, formula)
        ws.cell(row_idx, 3, f"=B{row_idx}/'05_원본근거'!$J$5")
        ws.cell(row_idx, 4, note)
        ws.cell(row_idx, 3).number_format = "0.0%"
        for col in range(1, 5):
            ws.cell(row_idx, col).font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])
            ws.cell(row_idx, col).alignment = Alignment(vertical="center", wrap_text=True)
    ws.conditional_formatting.add("B11:B17", DataBarRule(start_type="num", start_value=0, end_type="max", color=COLORS["cyan"]))

    ws["F9"] = "제안 규모"
    ws["F9"].font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["navy"])
    metrics = [
        ("추가 길/데크 셀", '=SUMIF(\'03_배치목록\'!$C$6:$C$23,"길",\'03_배치목록\'!$F$6:$F$23)'),
        ("추가 오브젝트", '=SUMIF(\'03_배치목록\'!$C$6:$C$23,"오브젝트",\'03_배치목록\'!$F$6:$F$23)'),
        ("P1 항목", '=COUNTIF(\'03_배치목록\'!$A$6:$A$23,"P1")'),
        ("미정 항목", '=COUNTIF(\'03_배치목록\'!$I$6:$I$23,"미정")'),
    ]
    for idx, (label, formula) in enumerate(metrics, start=10):
        ws.merge_cells(start_row=idx, start_column=6, end_row=idx, end_column=8)
        ws.merge_cells(start_row=idx, start_column=9, end_row=idx, end_column=12)
        ws.cell(idx, 6, label)
        ws.cell(idx, 9, formula)
        ws.cell(idx, 6).fill = fill(COLORS["navy2"])
        ws.cell(idx, 6).font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["white"])
        ws.cell(idx, 9).fill = fill(COLORS["pale_orange"])
        ws.cell(idx, 9).font = Font(name="맑은 고딕", size=14, bold=True, color=COLORS["navy"])
        ws.cell(idx, 6).alignment = ws.cell(idx, 9).alignment = Alignment(horizontal="center", vertical="center")

    findings = [
        ("1", "구조", "현재 길은 한 방향 ㄷ자 골격이라 우회·순환 선택지가 없음."),
        ("2", "밀도", "170개 Props 중 109개가 배경이고, 체감 장면을 만드는 비배경은 61개."),
        ("3", "반복", "수조 14개·부산물 13개·포스트 12개가 반복되어 개별 장면 구분이 약함."),
        ("4", "해법", "남측 광장 + 동측 세로축 + 중앙 연결로가 가장 적은 추가로 가장 큰 동선 변화를 만듦."),
        ("5", "배치", "큰 오브젝트는 어깨/모서리에, 중앙 3칸과 교차점 4칸은 항상 비워야 함."),
    ]
    ws.merge_cells("A20:L20")
    ws["A20"] = "핵심 판단"
    ws["A20"].fill = fill(COLORS["navy"])
    ws["A20"].font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["white"])
    ws["A20"].alignment = Alignment(horizontal="center")
    for row_idx, (no, category, text) in enumerate(findings, start=21):
        ws.cell(row_idx, 1, no)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        ws.cell(row_idx, 2, category)
        ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=12)
        ws.cell(row_idx, 4, text)
        ws.cell(row_idx, 1).fill = fill(COLORS["proposed"])
        ws.cell(row_idx, 1).font = Font(name="맑은 고딕", bold=True, color=COLORS["white"])
        ws.cell(row_idx, 2).fill = fill(COLORS["panel"])
        ws.cell(row_idx, 2).font = Font(name="맑은 고딕", bold=True, color=COLORS["navy"])
        ws.cell(row_idx, 4).fill = fill(COLORS["panel2"])
        ws.cell(row_idx, 4).font = Font(name="맑은 고딕", size=10, color=COLORS["ink"])
        for col in (1, 2, 4):
            ws.cell(row_idx, col).alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if col < 4 else "left")
        ws.row_dimensions[row_idx].height = 32

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["D"].width = 30
    ws.freeze_panes = "A9"
    ws.print_area = "A1:L26"
    ws.page_setup.orientation = "landscape"


def build_source(wb: Workbook) -> None:
    ws = wb.create_sheet("05_원본근거")
    ws.sheet_properties.tabColor = COLORS["muted"]
    set_sheet_defaults(ws)
    apply_title(
        ws,
        "원본 씬 근거 · 그룹 수량과 도로 좌표",
        f"읽기 전용 분석 대상: {SOURCE_SCENE} | SHA-256: {SOURCE_SHA256}",
        12,
    )

    headers = ["오브젝트 그룹", "개수", "Min X", "Max X", "Min Z", "Max Z", "역할"]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    style_header(ws[5][:7])
    for row_idx, values in enumerate(OBJECT_GROUPS, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
            ws.cell(row_idx, col_idx).font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])
        for col in range(3, 7):
            ws.cell(row_idx, col).number_format = "0.000"
    add_table(ws, f"A5:G{5+len(OBJECT_GROUPS)}", "CurrentObjectGroups", "TableStyleMedium2")

    summary_items = [
        ("현재 도로", len(ROAD_ROWS)),
        ("현재 Props", sum(row[1] for row in OBJECT_GROUPS)),
        ("해수 배경", OBJECT_GROUPS[0][1]),
        ("우천 바닥", OBJECT_GROUPS[3][1]),
        ("실질 비배경", "=J5-J6-J7"),
        ("도로 최소 X", "=MIN(B26:B46)"),
        ("도로 최대 X", "=MAX(B26:B46)"),
        ("도로 최소 Z", "=MIN(C26:C46)"),
        ("도로 최대 Z", "=MAX(C26:C46)"),
    ]
    for row_idx, (label, value) in enumerate(summary_items, start=4):
        ws.cell(row_idx, 9, label)
        ws.cell(row_idx, 10, value)
        ws.cell(row_idx, 9).fill = fill(COLORS["panel"])
        ws.cell(row_idx, 9).font = Font(name="맑은 고딕", size=9, bold=True, color=COLORS["navy"])
        ws.cell(row_idx, 10).fill = fill(COLORS["panel2"])
        ws.cell(row_idx, 10).font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])

    road_start = 25
    road_headers = ["도로 유형", "World X", "World Z", "Yaw", "구간", "소스", "비고"]
    for col, header in enumerate(road_headers, start=1):
        ws.cell(road_start, col, header)
    style_header(ws[road_start][:7])
    for row_idx, (road_type, x, z, yaw, segment) in enumerate(ROAD_ROWS, start=road_start + 1):
        values = (road_type, x, z, yaw, segment, "SM_Pier_Long_Fantasy", "저장 씬 Transform 기준")
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
            ws.cell(row_idx, col_idx).font = Font(name="맑은 고딕", size=9, color=COLORS["ink"])
        ws.cell(row_idx, 2).number_format = "0.000"
        ws.cell(row_idx, 3).number_format = "0.000"
        ws.cell(row_idx, 4).number_format = "0"
    road_end = road_start + len(ROAD_ROWS)
    add_table(ws, f"A{road_start}:G{road_end}", "CurrentRoadTransforms", "TableStyleMedium9")

    widths = [28, 11, 11, 11, 24, 28, 24, 3, 20, 18, 3, 3]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws["A3"].comment = Comment(f"분석 대상: {SOURCE_SCENE}\n파일 해시: {SOURCE_SHA256}", "User")
    ws.freeze_panes = "A6"
    ws.print_area = f"A1:J{road_end}"
    ws.page_setup.orientation = "landscape"


def rgb_from_cell(cell) -> tuple[int, int, int]:
    rgb = cell.fill.fgColor.rgb
    if not rgb or rgb in {"00000000", "FFFFFFFF"}:
        return (255, 255, 255)
    rgb = rgb[-6:]
    try:
        return tuple(int(rgb[i : i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return (255, 255, 255)


def font_path() -> str:
    candidates = [
        Path("C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/malgunbd.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    raise FileNotFoundError("No suitable font found")


def render_sheet(ws, cell_range: str, output_path: Path, paste_concept: bool = False) -> None:
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    widths = []
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        widths.append(max(42, int((ws.column_dimensions[letter].width or 10) * 7.2)))
    heights = [max(24, int((ws.row_dimensions[row].height or 18) * 1.33)) for row in range(min_row, max_row + 1)]
    x_positions = [0]
    y_positions = [0]
    for width in widths:
        x_positions.append(x_positions[-1] + width)
    for height in heights:
        y_positions.append(y_positions[-1] + height)

    extra_height = 0
    concept = None
    if paste_concept:
        concept = Image.open(CONCEPT_PATH).convert("RGB")
        concept.thumbnail((x_positions[-1], 680))
        extra_height = concept.height + 20
    canvas = Image.new("RGB", (x_positions[-1], y_positions[-1] + extra_height), "white")
    draw = ImageDraw.Draw(canvas)
    regular = ImageFont.truetype(font_path(), 16)
    small = ImageFont.truetype(font_path(), 13)
    bold = ImageFont.truetype("C:/Windows/Fonts/malgunbd.ttf" if Path("C:/Windows/Fonts/malgunbd.ttf").exists() else font_path(), 17)

    merged_lookup = {}
    for merged in ws.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_lookup[(row, col)] = merged

    handled = set()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            merged = merged_lookup.get((row, col))
            if merged:
                key = str(merged)
                if key in handled or row != merged.min_row or col != merged.min_col:
                    continue
                handled.add(key)
                c1 = max(merged.min_col, min_col)
                c2 = min(merged.max_col, max_col)
                r1 = max(merged.min_row, min_row)
                r2 = min(merged.max_row, max_row)
                x1 = x_positions[c1 - min_col]
                x2 = x_positions[c2 - min_col + 1]
                y1 = y_positions[r1 - min_row]
                y2 = y_positions[r2 - min_row + 1]
                cell = ws.cell(merged.min_row, merged.min_col)
            else:
                x1 = x_positions[col - min_col]
                x2 = x_positions[col - min_col + 1]
                y1 = y_positions[row - min_row]
                y2 = y_positions[row - min_row + 1]
                cell = ws.cell(row, col)
            draw.rectangle((x1, y1, x2, y2), fill=rgb_from_cell(cell), outline=(203, 213, 225), width=1)
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if text.startswith("="):
                text = text[1:]
            max_chars = max(4, int((x2 - x1) / 9))
            words = []
            line = ""
            for char in text:
                if len(line) >= max_chars and char not in " ,./-":
                    words.append(line)
                    line = char
                else:
                    line += char
            if line:
                words.append(line)
            text = "\n".join(words[:4])
            selected_font = bold if cell.font.bold else (small if x2 - x1 < 85 else regular)
            color = cell.font.color.rgb[-6:] if cell.font.color and cell.font.color.type == "rgb" and cell.font.color.rgb else "17202A"
            try:
                text_color = tuple(int(color[i : i + 2], 16) for i in (0, 2, 4))
            except ValueError:
                text_color = (23, 32, 42)
            box = draw.multiline_textbbox((0, 0), text, font=selected_font, spacing=2, align="center")
            tw, th = box[2] - box[0], box[3] - box[1]
            draw.multiline_text(((x1 + x2 - tw) / 2, (y1 + y2 - th) / 2), text, font=selected_font, fill=text_color, spacing=2, align="center")

    if concept is not None:
        x = (canvas.width - concept.width) // 2
        canvas.paste(concept, (x, y_positions[-1] + 10))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, quality=95)


def create_contact_sheet(preview_paths: list[Path], output_path: Path) -> None:
    thumbs = []
    for path in preview_paths:
        img = Image.open(path).convert("RGB")
        img.thumbnail((1050, 620))
        thumbs.append((path.stem, img.copy()))
    card_w, card_h = 1100, 700
    canvas = Image.new("RGB", (card_w * 2, card_h * 3), (236, 243, 245))
    draw = ImageDraw.Draw(canvas)
    title_font = ImageFont.truetype("C:/Windows/Fonts/malgunbd.ttf" if Path("C:/Windows/Fonts/malgunbd.ttf").exists() else font_path(), 24)
    for idx, (name, img) in enumerate(thumbs):
        x0 = (idx % 2) * card_w
        y0 = (idx // 2) * card_h
        draw.rectangle((x0 + 16, y0 + 16, x0 + card_w - 16, y0 + card_h - 16), fill="white", outline=(30, 58, 76), width=2)
        draw.text((x0 + 30, y0 + 28), name, fill=(19, 42, 58), font=title_font)
        canvas.paste(img, (x0 + (card_w - img.width) // 2, y0 + 70))
    canvas.save(output_path, quality=95)


def validate_workbook(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    expected = ["00_읽는법", "01_현재맵", "02_추가제안", "03_배치목록", "04_분석요약", "05_원본근거"]
    assert wb.sheetnames == expected, wb.sheetnames
    assert wb["05_원본근거"]["J4"].value == 21
    assert wb["05_원본근거"]["J5"].value == 170
    assert wb["00_읽는법"]["A1"].value.startswith("노량진 맵툴1")
    assert wb["03_배치목록"].max_row >= 23
    formula_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(token in value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}: {value}")
    assert not formula_errors, formula_errors


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    QA_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    build_readme(wb)
    build_current_map(wb)
    build_proposal_map(wb)
    build_placements(wb)
    build_summary(wb)
    build_source(wb)
    wb.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)

    rendered_wb = load_workbook(OUTPUT_PATH, data_only=False)
    preview_specs = [
        ("00_읽는법", "A1:P11", True),
        ("01_현재맵", "A1:W23", False),
        ("02_추가제안", "A1:W23", False),
        ("03_배치목록", "A1:J23", False),
        ("04_분석요약", "A1:L26", False),
        ("05_원본근거", "A1:J46", False),
    ]
    preview_paths = []
    for sheet_name, cell_range, paste_concept in preview_specs:
        path = QA_DIR / f"{sheet_name}.png"
        render_sheet(rendered_wb[sheet_name], cell_range, path, paste_concept=paste_concept)
        preview_paths.append(path)
    create_contact_sheet(preview_paths, QA_DIR / "all_sheets_contact.png")
    print(f"Workbook: {OUTPUT_PATH}")
    print(f"Sheets: {', '.join(rendered_wb.sheetnames)}")
    print("Validation: PASS")


if __name__ == "__main__":
    main()
