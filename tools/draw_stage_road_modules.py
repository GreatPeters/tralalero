from __future__ import annotations

import math
import random
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageDraw, ImageFilter

from rebuild_meshy_design_assets import DESIGN_DIR, IMAGE_DIR, kind_code, stage_code, target_filename


WIDTH = 1536
HEIGHT = 1024
SIZE = (WIDTH, HEIGHT)
BACKGROUND = (248, 248, 246, 255)

STYLE = {
    "STAGE01_NRY": {
        "top": (142, 90, 45),
        "top_light": (214, 148, 82),
        "top_dark": (79, 47, 27),
        "side": (25, 94, 137),
        "rim": (94, 60, 32),
        "seam": (63, 38, 23),
        "accent": (23, 107, 162),
        "metal": (62, 64, 68),
        "highlight": (255, 229, 183),
    },
    "STAGE02_HWY": {
        "top": (42, 47, 50),
        "top_light": (82, 88, 91),
        "top_dark": (18, 22, 25),
        "side": (118, 125, 123),
        "rim": (24, 29, 32),
        "seam": (31, 35, 37),
        "accent": (62, 68, 71),
        "metal": (155, 160, 158),
        "highlight": (136, 143, 145),
    },
    "STAGE03_RST": {
        "top": (117, 127, 121),
        "top_light": (175, 184, 174),
        "top_dark": (71, 83, 77),
        "side": (72, 91, 82),
        "rim": (196, 205, 188),
        "seam": (82, 92, 86),
        "accent": (134, 146, 138),
        "metal": (153, 162, 154),
        "highlight": (220, 225, 212),
    },
    "STAGE04_CITY": {
        "top": (71, 78, 88),
        "top_light": (118, 128, 137),
        "top_dark": (38, 44, 51),
        "side": (102, 110, 114),
        "rim": (180, 189, 190),
        "seam": (47, 53, 59),
        "accent": (92, 101, 110),
        "metal": (38, 42, 47),
        "highlight": (190, 198, 201),
    },
    "STAGE05_GNG": {
        "top": (31, 33, 38),
        "top_light": (71, 73, 80),
        "top_dark": (11, 13, 17),
        "side": (45, 48, 54),
        "rim": (15, 17, 21),
        "seam": (22, 24, 29),
        "accent": (61, 63, 70),
        "metal": (90, 92, 98),
        "highlight": (132, 134, 140),
    },
}


def rgba(color: tuple[int, ...], alpha: int = 255) -> tuple[int, int, int, int]:
    return (color[0], color[1], color[2], alpha)


def blend(a: tuple[int, int, int], b: tuple[int, int, int], t: float) -> tuple[int, int, int]:
    return tuple(int(a[i] + (b[i] - a[i]) * t) for i in range(3))


def fill_mask(image: Image.Image, mask: Image.Image, color: tuple[int, ...]) -> None:
    image.paste(Image.new("RGBA", SIZE, color if len(color) == 4 else rgba(color)), (0, 0), mask)


def shift_mask(mask: Image.Image, dx: int, dy: int) -> Image.Image:
    shifted = Image.new("L", SIZE, 0)
    shifted.paste(mask, (dx, dy))
    return shifted


def draw_line(draw: ImageDraw.ImageDraw, points: list[tuple[int, int]], fill, width: int) -> None:
    draw.line(points, fill=fill, width=width, joint="curve")


def path_mask(points: list[tuple[int, int]], width: int) -> Image.Image:
    mask = Image.new("L", SIZE, 0)
    draw_line(ImageDraw.Draw(mask), points, 255, width)
    return mask


def polygon_mask(points: list[tuple[int, int]]) -> Image.Image:
    mask = Image.new("L", SIZE, 0)
    ImageDraw.Draw(mask).polygon(points, fill=255)
    return mask


def clipped(image: Image.Image, mask: Image.Image, draw_fn) -> None:
    layer = Image.new("RGBA", SIZE, (0, 0, 0, 0))
    draw_fn(ImageDraw.Draw(layer, "RGBA"))
    alpha = ImageChops.multiply(layer.getchannel("A"), mask)
    layer.putalpha(alpha)
    image.alpha_composite(layer)


def shape_from_name(name_en: str) -> str:
    lower = name_en.lower()
    if "left" in lower:
        return "left"
    if "right" in lower:
        return "right"
    if any(word in lower for word in ("narrow", "merge", "bottleneck")):
        return "narrow"
    if any(word in lower for word in ("split", "exit")):
        return "split"
    return "straight"


def shape_mask(shape: str) -> tuple[Image.Image, list[list[tuple[int, int]]], list[tuple[int, int]] | None]:
    if shape == "straight":
        polygon = [(478, 116), (1058, 116), (1260, 884), (276, 884)]
        return polygon_mask(polygon), [[(768, 874), (768, 126)]], polygon
    if shape == "narrow":
        polygon = [(602, 116), (934, 116), (1212, 884), (324, 884)]
        return polygon_mask(polygon), [[(768, 874), (768, 126)]], polygon
    if shape == "left":
        center = [(910, 870), (798, 695), (655, 444), (520, 126)]
        return path_mask(center, 360), [center], None
    if shape == "right":
        center = [(626, 870), (738, 695), (881, 444), (1016, 126)]
        return path_mask(center, 360), [center], None

    trunk = [(768, 884), (768, 560)]
    left = [(768, 560), (610, 382), (450, 126)]
    right = [(768, 560), (926, 382), (1086, 126)]
    mask = Image.new("L", SIZE, 0)
    draw = ImageDraw.Draw(mask)
    for points in (trunk, left, right):
        draw_line(draw, points, 255, 305)
    draw.ellipse((632, 430, 904, 704), fill=255)
    return mask, [trunk, left, right], None


def draw_gradient_surface(image: Image.Image, mask: Image.Image, style: dict[str, tuple[int, int, int]]) -> None:
    gradient_size = (192, 128)
    gradient = Image.new("RGBA", gradient_size, (0, 0, 0, 0))
    pixels = gradient.load()
    for y in range(gradient_size[1]):
        vertical = y / (gradient_size[1] - 1)
        for x in range(gradient_size[0]):
            diagonal = (x / (gradient_size[0] - 1)) * 0.32 + (1.0 - vertical) * 0.68
            base = blend(style["top_dark"], style["top_light"], min(max(diagonal, 0.0), 1.0))
            pixels[x, y] = rgba(base)
    gradient = gradient.resize(SIZE, Image.Resampling.BICUBIC)
    gradient.putalpha(mask)
    image.alpha_composite(gradient)


def draw_object_base(image: Image.Image, top_mask: Image.Image, style: dict[str, tuple[int, int, int]]) -> Image.Image:
    outer = top_mask.filter(ImageFilter.MaxFilter(53))
    side = shift_mask(outer, 0, 56)
    shadow = shift_mask(outer, 28, 78).filter(ImageFilter.GaussianBlur(26))
    fill_mask(image, shadow, (0, 0, 0, 76))
    fill_mask(image, side, rgba(style["side"]))
    fill_mask(image, outer, rgba(style["rim"]))
    draw_gradient_surface(image, top_mask, style)
    return outer


def add_edge_highlights(image: Image.Image, top_mask: Image.Image, outer: Image.Image) -> None:
    top_edge = ImageChops.subtract(outer, top_mask).filter(ImageFilter.GaussianBlur(1.4))
    highlight = shift_mask(top_edge, -8, -14)
    shade = shift_mask(top_edge, 12, 16)
    fill_mask(image, highlight, (255, 255, 255, 46))
    fill_mask(image, shade, (0, 0, 0, 46))


def draw_rivet(draw: ImageDraw.ImageDraw, x: float, y: float, radius: int, metal: tuple[int, int, int]) -> None:
    draw.ellipse((x - radius, y - radius, x + radius, y + radius), fill=rgba((28, 29, 31), 190))
    draw.ellipse((x - radius + 2, y - radius + 2, x + radius - 2, y + radius - 2), fill=rgba(metal, 225))
    draw.ellipse((x - radius + 4, y - radius + 3, x - radius + 9, y - radius + 8), fill=(255, 255, 255, 120))


def draw_perspective_board_grid(draw: ImageDraw.ImageDraw, style: dict[str, tuple[int, int, int]], rng: random.Random) -> None:
    for y in range(155, 875, 72):
        left_x = int(462 - (y - 155) * 0.27)
        right_x = int(1074 + (y - 155) * 0.27)
        draw.line([(left_x, y), (right_x, y)], fill=rgba(style["seam"], 170), width=5)
        draw.line([(left_x + 8, y - 4), (right_x - 8, y - 4)], fill=(255, 230, 170, 28), width=2)
    for x in (616, 768, 920):
        draw.line([(x, 128), (x, 896)], fill=rgba(style["seam"], 155), width=5)
        draw.line([(x - 4, 128), (x - 20, 896)], fill=(255, 230, 170, 22), width=2)

    for y in range(170, 855, 72):
        for x in (610, 760, 928):
            draw_rivet(draw, x + rng.randint(-8, 8), y + rng.randint(-4, 4), 7, style["metal"])


def draw_pavement_segments(draw: ImageDraw.ImageDraw, style: dict[str, tuple[int, int, int]], stage: str) -> None:
    if stage == "STAGE02_HWY":
        for y in range(210, 845, 118):
            draw.line([(410, y), (1128, y - 15)], fill=rgba(style["seam"], 80), width=5)
        return
    if stage == "STAGE03_RST":
        for y in range(190, 850, 132):
            draw.line([(370, y), (1170, y)], fill=rgba(style["seam"], 88), width=5)
        return
    if stage == "STAGE04_CITY":
        for y in range(205, 850, 122):
            draw.line([(380, y), (1160, y)], fill=rgba(style["seam"], 82), width=5)
        for x in (610, 928):
            draw.line([(x, 145), (x, 870)], fill=rgba(style["seam"], 58), width=4)
        return
    if stage == "STAGE05_GNG":
        for y in range(230, 835, 150):
            draw.line([(470, y), (1066, y - 22)], fill=(255, 255, 255, 26), width=5)


def draw_noise_texture(
    draw: ImageDraw.ImageDraw,
    rng: random.Random,
    style: dict[str, tuple[int, int, int]],
    count: int,
    alpha: tuple[int, int],
    radius: tuple[int, int],
) -> None:
    for _ in range(count):
        x = rng.randint(330, 1190)
        y = rng.randint(140, 860)
        w = rng.randint(radius[0], radius[1])
        h = rng.randint(max(4, radius[0] // 3), max(8, radius[1] // 2))
        tone = rng.choice([style["top_light"], style["top_dark"], style["accent"]])
        draw.ellipse((x, y, x + w, y + h), fill=rgba(tone, rng.randint(alpha[0], alpha[1])))


def draw_water_highlights(draw: ImageDraw.ImageDraw, rng: random.Random) -> None:
    for _ in range(20):
        x = rng.randint(360, 1100)
        y = rng.randint(155, 850)
        w = rng.randint(65, 180)
        h = rng.randint(15, 48)
        draw.ellipse((x, y, x + w, y + h), fill=(40, 133, 190, rng.randint(100, 150)))
        draw.arc((x + 5, y + 3, x + w - 5, y + h - 3), 190, 350, fill=(220, 247, 255, 92), width=3)
    for _ in range(24):
        x = rng.randint(345, 1120)
        y = rng.randint(170, 860)
        draw.line([(x, y), (x + rng.randint(24, 72), y - rng.randint(3, 12))], fill=(255, 255, 255, rng.randint(38, 78)), width=2)


def draw_wood_grain(draw: ImageDraw.ImageDraw, rng: random.Random, style: dict[str, tuple[int, int, int]]) -> None:
    for _ in range(135):
        x = rng.randint(360, 1110)
        y = rng.randint(145, 870)
        length = rng.randint(35, 135)
        curve = rng.randint(-8, 8)
        color = rgba(blend(style["top_dark"], style["top"], rng.random() * 0.5), rng.randint(42, 96))
        draw.line([(x, y), (x + length // 2, y + curve), (x + length, y + rng.randint(-5, 5))], fill=color, width=rng.randint(1, 3))


def draw_asphalt_texture(draw: ImageDraw.ImageDraw, rng: random.Random, style: dict[str, tuple[int, int, int]], dense: bool = True) -> None:
    for _ in range(560 if dense else 320):
        x = rng.randint(320, 1210)
        y = rng.randint(130, 880)
        s = rng.randint(1, 4)
        color = rng.choice([style["top_dark"], style["top_light"], style["accent"]])
        draw.ellipse((x, y, x + s, y + s), fill=rgba(color, rng.randint(34, 95)))
    draw_noise_texture(draw, rng, style, 28 if dense else 18, (22, 54), (32, 140))


def draw_gloss(draw: ImageDraw.ImageDraw, rng: random.Random, style: dict[str, tuple[int, int, int]]) -> None:
    for _ in range(24):
        x = rng.randint(430, 1020)
        y = rng.randint(145, 810)
        w = rng.randint(90, 260)
        draw.line([(x, y), (x + w, y - rng.randint(8, 34))], fill=(255, 255, 255, rng.randint(22, 52)), width=rng.randint(3, 8))
    draw_noise_texture(draw, rng, style, 16, (18, 42), (55, 170))


def draw_stage_surface(
    draw: ImageDraw.ImageDraw,
    stage: str,
    shape: str,
    style: dict[str, tuple[int, int, int]],
    rng: random.Random,
) -> None:
    if stage == "STAGE01_NRY":
        draw_perspective_board_grid(draw, style, rng)
        draw_wood_grain(draw, rng, style)
        draw_water_highlights(draw, rng)
        if shape in ("straight", "narrow"):
            for x, y in ((372, 770), (1084, 770), (462, 155), (992, 155)):
                draw.rounded_rectangle((x, y, x + 88, y + 60), radius=10, fill=rgba(style["metal"], 230))
                draw_rivet(draw, x + 22, y + 18, 7, (205, 201, 188))
                draw_rivet(draw, x + 65, y + 18, 7, (205, 201, 188))
        return

    if stage == "STAGE02_HWY":
        draw_pavement_segments(draw, style, stage)
        draw_asphalt_texture(draw, rng, style)
        return

    if stage == "STAGE03_RST":
        draw_pavement_segments(draw, style, stage)
        draw_asphalt_texture(draw, rng, style, dense=False)
        return

    if stage == "STAGE04_CITY":
        draw_pavement_segments(draw, style, stage)
        draw_asphalt_texture(draw, rng, style)
        return

    if stage == "STAGE05_GNG":
        draw_gloss(draw, rng, style)
        return


def draw_road_module(asset_id: str, stage: str, name_en: str, output: Path) -> None:
    style = STYLE[stage]
    shape = shape_from_name(name_en)
    top_mask, _, polygon = shape_mask(shape)
    rng = random.Random(f"{asset_id}:{name_en}:v3")

    image = Image.new("RGBA", SIZE, BACKGROUND)
    outer = draw_object_base(image, top_mask, style)
    add_edge_highlights(image, top_mask, outer)

    def details(draw: ImageDraw.ImageDraw) -> None:
        draw_stage_surface(draw, stage, shape, style, rng)
        if polygon is not None:
            draw.line(polygon + [polygon[0]], fill=rgba(style["seam"], 150), width=4)
            draw.line([(polygon[0][0] + 20, polygon[0][1] + 20), (polygon[1][0] - 20, polygon[1][1] + 20)], fill=(255, 255, 255, 34), width=3)

    clipped(image, top_mask, details)
    image = image.filter(ImageFilter.UnsharpMask(radius=1.2, percent=115, threshold=3))
    output.parent.mkdir(parents=True, exist_ok=True)
    image.convert("RGB").save(output, quality=95)


def iter_road_rows():
    workbook = load_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", read_only=True, data_only=True)
    worksheet = workbook.worksheets[1]
    for sequence, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        asset_id = str(row[0]) if row[0] else ""
        stage = stage_code(asset_id, row[1])
        if stage not in STYLE or kind_code(asset_id, row[4]) != "ROAD":
            continue
        yield asset_id, stage, str(row[3]), IMAGE_DIR / target_filename(sequence, row)


def main() -> None:
    generated = []
    for asset_id, stage, name_en, output in iter_road_rows():
        draw_road_module(asset_id, stage, name_en, output)
        generated.append(output)

    for output in generated:
        print(output.name)
    print(f"Generated {len(generated)} Meshy-ready stage road image(s).")


if __name__ == "__main__":
    main()
