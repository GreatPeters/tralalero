from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "design"
OUT_FILE = OUT_DIR / "트랄랄레오_MeshyAI_소품리스트_한글.xlsx"


def style_sheet_headers(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B7C3D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border


def style_sheet_body(ws):
    thin = Side(style="thin", color="B7C3D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
            cell.border = border


def build_workbook():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    route_ws = wb.active
    route_ws.title = "구간표"
    asset_ws = wb.create_sheet("소품리스트")
    prompt_ws = wb.create_sheet("프롬프트")

    route_headers = [
        "구간",
        "예상 시간",
        "지역 컨셉",
        "화면에서 먼저 읽혀야 하는 것",
        "플레이 포인트",
        "다음 구간으로 넘어가는 신호",
    ]
    route_rows = [
        [
            "노량진 1",
            "0:00 ~ 1:10",
            "부두에서 수산시장 초입으로 진입",
            "젖은 나무 바닥, 밧줄, 파란 박스, 갈매기",
            "러너 기본 리듬 소개, 쉬운 장애물, 첫 물고기 픽업",
            "생선 박스와 수조가 늘어나고 육지 쪽 바닥 비율이 커짐",
        ],
        [
            "노량진 2",
            "1:10 ~ 2:10",
            "붐비는 수산시장 길목",
            "스티로폼 박스, 수조, 손수레, 물웅덩이",
            "좁은 길 판독, 갑작스런 꺾임, 첫 이동 장애물",
            "냉동차, 진입 램프, 도로 구조물이 보여 고속도로를 예고",
        ],
        [
            "고속도로",
            "2:10 ~ 4:10",
            "외곽 순환 고속도로",
            "차선, 가드레일, 표지판, 공사 배리케이드",
            "긴 직선 구간, 속도감, 좌우 선택 압박",
            "멀리 휴게소 간판과 주차장 조명이 보이기 시작",
        ],
        [
            "휴게소",
            "4:10 ~ 5:20",
            "잠깐 숨 돌리는 서비스 에어리어",
            "자판기, 주차칸, 편의점 간판, 쓰레기통",
            "템포 완급 조절, 짧은 장애물 묶음, 보상 루트",
            "도시 스카이라인이 커지고 출구가 도심 쪽으로 좁아짐",
        ],
        [
            "도시",
            "5:20 ~ 7:50",
            "활발한 도심 진입",
            "버스정류장, 공사 펜스, 신호등, 간판",
            "점프와 회피가 함께 필요한 밀도 높은 구간",
            "신발 광고판, 유리 쇼윈도, 고급 상권 분위기가 점점 강해짐",
        ],
        [
            "강남",
            "7:50 ~ 10:00",
            "고급 상권과 백화점 도착",
            "쇼윈도, 부티크 간판, 백화점 배너, 진열대",
            "최종 압박, 가장 빡센 적 조합, 목적지 도달 연출",
            "백화점 입구와 딱 맞는 신발 발견으로 마무리",
        ],
    ]
    route_ws.append(route_headers)
    for row in route_rows:
        route_ws.append(row)
    style_sheet_headers(route_ws)
    style_sheet_body(route_ws)
    for i, width in enumerate([16, 14, 28, 34, 34, 40], start=1):
        route_ws.column_dimensions[get_column_letter(i)].width = width
    route_ws.freeze_panes = "A2"

    asset_headers = [
        "소품 코드",
        "구간",
        "소품명",
        "분류",
        "용도",
        "우선순위",
        "Meshy 적합도",
        "추천 방식",
        "재사용 방식",
        "난이도 재활용",
        "비주얼 메모",
        "ChatGPT에 넣을 짧은 요청문",
    ]
    asset_rows = [
        [
            "NRY-001",
            "노량진",
            "파란 플라스틱 생선 박스",
            "장식물",
            "길 옆 장식",
            "P0",
            "높음",
            "Image to 3D",
            "색만 바꿔서 여러 개 반복",
            "노말/나이트메어/하드 공통",
            "두꺼운 플라스틱 느낌, 위쪽에 생선이나 얼음이 살짝 보이면 좋음",
            "노량진 수산시장 세계관의 파란 플라스틱 생선 박스를 하나만 크게 그려줘. 흰 배경, 3/4 시점, 모바일 게임용으로 실루엣이 분명하게.",
        ],
        [
            "NRY-002",
            "노량진",
            "스티로폼 생선 상자",
            "장식물",
            "길 옆 장식",
            "P0",
            "높음",
            "Image to 3D",
            "단일형과 적층형 둘 다 활용",
            "노말/나이트메어/하드 공통",
            "두꺼운 스티로폼 벽, 생선 꼬리나 물기 표현은 단순하게",
            "노량진 수산시장 스타일의 스티로폼 생선 상자를 하나만 그려줘. 흰 배경, 단순한 형태, Meshy용으로 읽기 쉽게.",
        ],
        [
            "NRY-003",
            "노량진",
            "얼음 수조",
            "장식물",
            "지역 대표 랜드마크",
            "P0",
            "높음",
            "Image to 3D",
            "크기 2종으로 반복",
            "노말/나이트메어/하드 공통",
            "유리 프레임은 두껍게, 얼음과 해산물은 과하지 않게",
            "수산시장용 얼음 수조를 하나만 그려줘. 유리 프레임이 분명하고 모바일 게임 소품처럼 단순화된 형태로.",
        ],
        [
            "NRY-004",
            "노량진",
            "수산시장 손수레",
            "장애물",
            "이동 장애물 후보",
            "P0",
            "중간",
            "Image to 3D",
            "고급형 1개, 단순형 1개",
            "나이트메어/하드에서 더 자주 사용",
            "바퀴와 몸통이 크고 단순해야 함, 손잡이는 너무 얇지 않게",
            "수산시장 손수레를 하나만 그려줘. 모바일 러너 게임 장애물처럼 두껍고 읽기 쉽게.",
        ],
        [
            "NRY-005",
            "노량진",
            "미끄럼 경고 콘",
            "장애물",
            "정적 장애물",
            "P0",
            "높음",
            "Text to 3D",
            "색상만 바꿔 반복",
            "노말/나이트메어/하드 공통",
            "밑판이 넓고, 경고색 대비가 강해야 함",
            "젖은 바닥 경고 콘을 하나만 그려줘. 수산시장 바닥에 어울리게 하고 실루엣이 분명하게.",
        ],
        [
            "NRY-009",
            "노량진",
            "떨어진 물고기 픽업",
            "픽업",
            "재화 획득",
            "P0",
            "높음",
            "Image to 3D",
            "어종 3종, 크기 2종",
            "노말/나이트메어/하드 공통",
            "작아도 눈에 띄게 색 대비와 반짝임 느낌이 있어야 함",
            "줍는 보상 오브젝트처럼 보이는 떨어진 물고기를 하나만 그려줘. 흰 배경, 모바일 화면에서 잘 보이게.",
        ],
        [
            "HWY-002",
            "고속도로",
            "공사 배리케이드",
            "장애물",
            "정적 장애물",
            "P0",
            "높음",
            "Image to 3D",
            "색상 변형 2종",
            "노말/나이트메어/하드 공통",
            "X 표시와 받침이 크게 보여야 함",
            "고속도로 공사 배리케이드를 하나만 그려줘. 장애물로 바로 읽히게 강한 실루엣으로.",
        ],
        [
            "HWY-004",
            "고속도로",
            "고속도로 표지판",
            "장식물",
            "지역 인지 랜드마크",
            "P0",
            "중간",
            "Image to 3D",
            "지명만 바꿔 재사용",
            "노말/나이트메어/하드 공통",
            "오버헤드 표지판 프레임을 크게",
            "한국 고속도로 표지판을 하나만 그려줘. 흰 배경, 단순한 3D 게임 소품처럼.",
        ],
        [
            "RST-001",
            "휴게소",
            "자판기",
            "장식물",
            "지역 대표 소품",
            "P0",
            "높음",
            "Image to 3D",
            "음료형/간식형 변형",
            "노말/나이트메어/하드 공통",
            "정면 패널이 밝고 큼직해야 함",
            "한국 휴게소 자판기를 하나만 그려줘. 흰 배경, 읽기 쉬운 3D 게임 소품 스타일로.",
        ],
        [
            "CITY-001",
            "도시",
            "버스정류장 쉘터",
            "장식물",
            "도심 대표 랜드마크",
            "P0",
            "중간",
            "Image to 3D",
            "광고만 바꿔 반복",
            "노말/나이트메어/하드 공통",
            "지붕, 기둥, 벤치를 크게 단순화",
            "도시 버스정류장 쉘터를 하나만 그려줘. 한국 도심 느낌, 흰 배경, 단순한 3D 게임 소품으로.",
        ],
        [
            "CITY-002",
            "도시",
            "공사 안전 펜스",
            "장애물",
            "정적 장애물",
            "P0",
            "높음",
            "Image to 3D",
            "길이 2종",
            "노말/나이트메어/하드 공통",
            "메시망 느낌은 줄이고 프레임을 두껍게",
            "공사 안전 펜스를 하나만 그려줘. 장애물로 읽히게 강한 형태로.",
        ],
        [
            "CITY-005",
            "도시",
            "신발 광고 빌보드",
            "장식물",
            "스토리 예고 랜드마크",
            "P0",
            "중간",
            "Image to 3D",
            "광고 그래픽만 변경",
            "노말/나이트메어/하드 공통",
            "프레임은 크고 단순하게, 광고 면적 넓게",
            "신발 광고 빌보드 프레임을 하나만 그려줘. 흰 배경, 도시 배경 소품처럼.",
        ],
        [
            "GNG-001",
            "강남",
            "쇼윈도 신발 진열대",
            "장식물",
            "목적지 보상 연출",
            "P0",
            "높음",
            "Image to 3D",
            "높이 3종 변형",
            "노말/나이트메어/하드 공통",
            "고급 받침대 형태, 신발 실루엣은 하나만 크게",
            "고급 쇼윈도용 신발 진열대를 하나만 그려줘. 강남 상권 느낌으로 세련되지만 단순하게.",
        ],
        [
            "GNG-003",
            "강남",
            "부티크 간판",
            "장식물",
            "고급 상권 인지",
            "P0",
            "중간",
            "Image to 3D",
            "상호만 바꿔 재사용",
            "노말/나이트메어/하드 공통",
            "유리/금속 느낌이 나되 구조는 단순하게",
            "고급 부티크 간판을 하나만 그려줘. 강남 쇼핑 거리 세계관에 맞게.",
        ],
        [
            "GNG-007",
            "강남",
            "백화점 배너 기둥",
            "장식물",
            "최종 목적지 표시",
            "P0",
            "중간",
            "Image to 3D",
            "배너만 바꿔 반복",
            "노말/나이트메어/하드 공통",
            "기둥은 높지만 두껍게, 배너 판은 크게",
            "백화점 배너 기둥을 하나만 그려줘. 최종 목적지 느낌이 나게 고급스럽게.",
        ],
    ]
    asset_ws.append(asset_headers)
    for row in asset_rows:
        asset_ws.append(row)
    style_sheet_headers(asset_ws)
    style_sheet_body(asset_ws)
    for row in asset_ws.iter_rows(min_row=2, max_row=asset_ws.max_row):
        priority_cell = row[5]
        fit_cell = row[6]
        if priority_cell.value == "P0":
            priority_cell.fill = PatternFill("solid", fgColor="FCE4D6")
        if fit_cell.value == "높음":
            fit_cell.fill = PatternFill("solid", fgColor="E2F0D9")
        elif fit_cell.value == "중간":
            fit_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for i, width in enumerate([12, 12, 22, 12, 16, 10, 12, 14, 18, 18, 34, 54], start=1):
        asset_ws.column_dimensions[get_column_letter(i)].width = width
    asset_ws.freeze_panes = "A2"

    prompt_headers = ["항목", "내용"]
    prompt_rows = [
        [
            "기본 프롬프트 템플릿",
            "첨부한 참고 이미지를 스타일 레퍼런스로 사용해줘. 같은 세계관의 stylized 3D 모바일 러너 게임 소품을 하나만 만들어줘. 배경은 흰색 또는 단색으로 두고, 정면 또는 3/4 시점으로, 형태가 두껍고 읽기 쉽게 표현해줘. MeshyAI의 Image to 3D로 넘기기 좋게 얇은 구조와 잔 디테일은 줄여줘.",
        ],
        ["노량진 스타일 키워드", "젖은 나무 바닥, 수산시장, 파란 박스, 스티로폼 상자, 수조, 갈매기, 밧줄, 항구, 활기찬 시장 분위기"],
        ["고속도로 스타일 키워드", "한국 고속도로, 차선, 가드레일, 공사 배리케이드, 표지판, 반사띠, 두꺼운 금속 구조"],
        ["휴게소 스타일 키워드", "한국 휴게소, 자판기, 편의점, 주차장, 간식, 쓰레기통, 밝은 간판"],
        ["도시 스타일 키워드", "한국 도심, 버스정류장, 신호등, 공사 펜스, 간판, 택시, 보행 환경"],
        ["강남 스타일 키워드", "고급 쇼핑 거리, 유리 쇼윈도, 부티크 간판, 백화점, 고급 진열대, 프리미엄 분위기"],
        ["장애물일 때 추가 문장", "플레이어가 부딪히면 위험해 보이도록, 막는 느낌이 강한 실루엣과 색 대비로 표현해줘."],
        ["픽업일 때 추가 문장", "작지만 모바일 화면에서 잘 보이고, 보상처럼 보이게 밝고 또렷하게 표현해줘."],
        ["장식물일 때 추가 문장", "배경 장식물처럼 풍부하게 보이되, 장애물로 오해되지 않게 공격적인 형태는 줄여줘."],
        ["Meshy용 주의사항", "한 번에 한 소품만 생성하고, 복잡한 배경, 얇은 선, 작은 글자, 과한 투명효과, 지나치게 겹치는 구조는 피하는 게 좋다."],
    ]
    prompt_ws.append(prompt_headers)
    for row in prompt_rows:
        prompt_ws.append(row)
    style_sheet_headers(prompt_ws)
    style_sheet_body(prompt_ws)
    prompt_ws.column_dimensions["A"].width = 22
    prompt_ws.column_dimensions["B"].width = 120
    prompt_ws.freeze_panes = "A2"

    wb.save(OUT_FILE)
    return OUT_FILE


if __name__ == "__main__":
    result = build_workbook()
    print(result)
