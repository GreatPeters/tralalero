from __future__ import annotations

import argparse
import json
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DESIGN_DIR = ROOT / "docs" / "design"
IMAGE_DIR = ROOT / "output" / "meshy_images"
RND_DIR = IMAGE_DIR / "RnD"

WORKBOOKS = [
    (DESIGN_DIR / "tralalero_meshy_asset_plan.xlsx", False),
    (DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", True),
]

NUMBERED_IMAGE = re.compile(r"^\d{3}_.+\.png$")
ACTIVE_IMAGE = re.compile(
    r"^(?P<sequence>\d{3})_(?P<stage>STAGE\d{2}_[A-Z]+|COMMON)_"
    r"(?P<kind>[A-Z]+)_(?P<asset_number>\d{3})_(?P<name>.+)\.png$"
)
AUDIT_ASSET_REF = re.compile(r"(?:\d{3}_)?([A-Z]+-\d{3})(?:_[A-Za-z0-9][A-Za-z0-9_-]*)?")


@dataclass(frozen=True)
class RoadAsset:
    asset_id: str
    stage_code: str
    source_rel: str
    name_en: str
    name_kr: str
    role_en: str
    role_kr: str
    visual_notes_en: str
    visual_notes_kr: str
    brief: str

    @property
    def asset_number(self) -> str:
        return self.asset_id.split("-")[1]

    @property
    def rest_name(self) -> str:
        return f"{self.stage_code}_ROAD_{self.asset_number}_{slug(self.name_en)}.png"

    @property
    def source_path(self) -> Path:
        return RND_DIR / self.source_rel


@dataclass(frozen=True)
class ActiveEntry:
    asset_id: str
    name_en: str
    rest_name: str
    old_filename: str | None = None
    road: RoadAsset | None = None

    def filename(self, sequence: int) -> str:
        return f"{sequence:03d}_{self.rest_name}"


def slug(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    value = re.sub(r"_+", "_", value)
    return value


def road_notes_en(theme: str) -> str:
    return (
        f"Road-only full 3D MeshyAI reference: {theme}, visible slab thickness, "
        "top 3/4 asset view, isolated on a clean background; no characters, vehicles, "
        "coins, buildings, or full environment scene."
    )


def road_notes_kr(theme: str) -> str:
    return (
        f"Road-only 풀 3D MeshyAI 기준 이미지: {theme}, 보이는 바닥 두께, "
        "3/4 탑뷰 에셋 프리뷰, 깨끗한 배경 분리. 캐릭터, 차량, 코인, 건물, 전체 배경 씬 없음."
    )


def brief(name: str, theme: str) -> str:
    return (
        f"Single {name}, road only, full 3D MeshyAI asset preview, {theme}, "
        "visible thickness, clean background, 3/4 top view."
    )


ROAD_ASSETS_BY_STAGE: dict[str, list[RoadAsset]] = {
    "STAGE01_NRY": [
        RoadAsset(
            "NRY-038",
            "STAGE01_NRY",
            "01_noryangjin/01_STAGE01_NRY_road_only_wet_straight_pier.png",
            "Noryangjin wet straight pier road module",
            "노량진 젖은 직선 부두 길 모듈",
            "Straight stage path",
            "직선 스테이지 길",
            road_notes_en("wet wooden pier deck, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 부두 데크, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin wet straight pier road module", "wet wooden planks, raised beams, dock posts, ropes, and puddle accents"),
        ),
        RoadAsset(
            "NRY-039",
            "STAGE01_NRY",
            "01_noryangjin/02_STAGE01_NRY_road_only_wet_s_curve_pier.png",
            "Noryangjin wet S curve pier road module",
            "노량진 젖은 S자 부두 길 모듈",
            "S-curve stage path",
            "S자 스테이지 길",
            road_notes_en("wet wooden S-curve pier deck, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 S자 부두 데크, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin wet S-curve pier road module", "wet wooden S-curve planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-040",
            "STAGE01_NRY",
            "01_noryangjin/03_STAGE01_NRY_road_only_wet_t_junction_pier.png",
            "Noryangjin wet T junction pier road module",
            "노량진 젖은 T자 부두 길 모듈",
            "T-junction stage path",
            "T자 스테이지 길",
            road_notes_en("wet wooden T-junction pier deck, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 T자 부두 데크, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin wet T-junction pier road module", "wet wooden T-junction planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-041",
            "STAGE01_NRY",
            "01_noryangjin/04_STAGE01_NRY_road_only_left_90_pier_corner.png",
            "Noryangjin left 90 pier corner road module",
            "노량진 왼쪽 90도 부두 코너 길 모듈",
            "Left 90-degree stage path",
            "왼쪽 90도 스테이지 길",
            road_notes_en("wet wooden left 90-degree pier corner, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 왼쪽 90도 부두 코너, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin left 90-degree pier corner road module", "wet wooden corner planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-042",
            "STAGE01_NRY",
            "01_noryangjin/05_STAGE01_NRY_road_only_right_90_pier_corner.png",
            "Noryangjin right 90 pier corner road module",
            "노량진 오른쪽 90도 부두 코너 길 모듈",
            "Right 90-degree stage path",
            "오른쪽 90도 스테이지 길",
            road_notes_en("wet wooden right 90-degree pier corner, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 오른쪽 90도 부두 코너, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin right 90-degree pier corner road module", "wet wooden corner planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-043",
            "STAGE01_NRY",
            "01_noryangjin/06_STAGE01_NRY_road_only_cross_pier_intersection.png",
            "Noryangjin cross pier intersection road module",
            "노량진 십자 부두 교차로 길 모듈",
            "Cross intersection stage path",
            "십자 교차 스테이지 길",
            road_notes_en("wet wooden cross-intersection pier deck, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 십자 부두 교차로 데크, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin cross pier intersection road module", "wet wooden cross-intersection planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-044",
            "STAGE01_NRY",
            "01_noryangjin/07_STAGE01_NRY_road_only_hairpin_pier_uturn.png",
            "Noryangjin hairpin pier U turn road module",
            "노량진 헤어핀 부두 유턴 길 모듈",
            "Hairpin U-turn stage path",
            "헤어핀 유턴 스테이지 길",
            road_notes_en("wet wooden hairpin pier U-turn, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 헤어핀 부두 유턴, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin hairpin pier U-turn road module", "wet wooden hairpin planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-045",
            "STAGE01_NRY",
            "01_noryangjin/08_STAGE01_NRY_road_only_narrowing_pier_connector.png",
            "Noryangjin narrowing pier connector road module",
            "노량진 좁아지는 부두 연결 길 모듈",
            "Narrowing stage path",
            "좁아지는 스테이지 길",
            road_notes_en("wet wooden narrowing pier connector, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 좁아지는 부두 연결로, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin narrowing pier connector road module", "wet wooden narrowing planks, raised beams, dock posts, ropes, and puddles"),
        ),
        RoadAsset(
            "NRY-046",
            "STAGE01_NRY",
            "01_noryangjin/09_STAGE01_NRY_road_only_y_split_pier_fork.png",
            "Noryangjin Y split pier fork road module",
            "노량진 Y자 부두 분기 길 모듈",
            "Y-split stage path",
            "Y자 분기 스테이지 길",
            road_notes_en("wet wooden Y-split pier fork, raised side beams, dock posts, ropes, rivet strips, and puddle accents"),
            road_notes_kr("젖은 목재 Y자 부두 분기, 입체 측면 보, 부두 말뚝, 로프, 리벳 스트립, 물웅덩이 포인트"),
            brief("Noryangjin Y-split pier fork road module", "wet wooden Y-split planks, raised beams, dock posts, ropes, and puddles"),
        ),
    ],
    "STAGE02_HWY": [
        RoadAsset("HWY-032", "STAGE02_HWY", "02_highway/01_STAGE02_HWY_road_only_straight_expressway.png", "Highway straight expressway road module", "고속도로 직선 본선 길 모듈", "Straight stage path", "직선 스테이지 길", road_notes_en("elevated expressway asphalt, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 고속도로 아스팔트, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway straight expressway road module", "elevated asphalt, guardrails, road shoulders, lane markings, and concrete slab edges")),
        RoadAsset("HWY-033", "STAGE02_HWY", "02_highway/02_STAGE02_HWY_road_only_elevated_curve_ramp.png", "Highway elevated curve ramp road module", "고속도로 입체 커브 램프 길 모듈", "Curved ramp stage path", "커브 램프 스테이지 길", road_notes_en("elevated curved expressway ramp, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 곡선 고속도로 램프, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway elevated curve ramp road module", "elevated curved asphalt ramp, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-034", "STAGE02_HWY", "02_highway/03_STAGE02_HWY_road_only_y_split_expressway.png", "Highway Y split expressway road module", "고속도로 Y자 분기 길 모듈", "Y-split stage path", "Y자 분기 스테이지 길", road_notes_en("elevated Y-split expressway, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 Y자 고속도로 분기, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway Y-split expressway road module", "elevated Y-split asphalt, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-035", "STAGE02_HWY", "02_highway/04_STAGE02_HWY_road_only_left_90_expressway_corner.png", "Highway left 90 expressway corner road module", "고속도로 왼쪽 90도 코너 길 모듈", "Left 90-degree stage path", "왼쪽 90도 스테이지 길", road_notes_en("elevated left 90-degree expressway corner, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 왼쪽 90도 고속도로 코너, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway left 90-degree expressway corner road module", "elevated left corner asphalt, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-036", "STAGE02_HWY", "02_highway/05_STAGE02_HWY_road_only_right_90_expressway_corner.png", "Highway right 90 expressway corner road module", "고속도로 오른쪽 90도 코너 길 모듈", "Right 90-degree stage path", "오른쪽 90도 스테이지 길", road_notes_en("elevated right 90-degree expressway corner, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 오른쪽 90도 고속도로 코너, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway right 90-degree expressway corner road module", "elevated right corner asphalt, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-037", "STAGE02_HWY", "02_highway/06_STAGE02_HWY_road_only_cross_expressway_interchange.png", "Highway cross expressway interchange road module", "고속도로 십자 인터체인지 길 모듈", "Cross interchange stage path", "십자 교차 스테이지 길", road_notes_en("elevated cross expressway interchange, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 십자 고속도로 인터체인지, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway cross expressway interchange road module", "elevated cross interchange asphalt, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-038", "STAGE02_HWY", "02_highway/07_STAGE02_HWY_road_only_elevated_hairpin_ramp.png", "Highway elevated hairpin ramp road module", "고속도로 입체 헤어핀 램프 길 모듈", "Hairpin ramp stage path", "헤어핀 램프 스테이지 길", road_notes_en("elevated hairpin expressway ramp, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 헤어핀 고속도로 램프, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway elevated hairpin ramp road module", "elevated hairpin asphalt ramp, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-039", "STAGE02_HWY", "02_highway/08_STAGE02_HWY_road_only_narrowing_merge_expressway.png", "Highway narrowing merge expressway road module", "고속도로 좁아지는 합류 길 모듈", "Narrowing merge stage path", "좁아지는 합류 스테이지 길", road_notes_en("elevated narrowing expressway merge, guardrails, road shoulders, lane markings, and concrete slab edges"), road_notes_kr("입체 좁아지는 고속도로 합류로, 가드레일, 갓길, 차선 표시, 콘크리트 슬래브 모서리"), brief("Highway narrowing merge expressway road module", "elevated narrowing asphalt merge, guardrails, shoulders, lane markings, and slab edges")),
        RoadAsset("HWY-040", "STAGE02_HWY", "02_highway/09_STAGE02_HWY_road_only_toll_lane_approach.png", "Highway toll lane approach road module", "고속도로 톨게이트 진입 길 모듈", "Toll lane stage path", "톨게이트 진입 스테이지 길", road_notes_en("elevated toll-lane approach asphalt, lane separators, guardrails, road shoulders, and concrete slab edges"), road_notes_kr("입체 톨게이트 진입 아스팔트, 차선 분리대, 가드레일, 갓길, 콘크리트 슬래브 모서리"), brief("Highway toll lane approach road module", "elevated toll approach asphalt, lane separators, guardrails, shoulders, and slab edges")),
    ],
    "STAGE03_RST": [
        RoadAsset("RST-024", "STAGE03_RST", "03_rest_stop/01_STAGE03_RST_road_only_straight_service_lane.png", "Rest stop straight service lane road module", "휴게소 직선 서비스 차로 길 모듈", "Straight stage path", "직선 스테이지 길", road_notes_en("service-lane pavement, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("휴게소 서비스 차로 포장, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop straight service lane road module", "service-lane pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-025", "STAGE03_RST", "03_rest_stop/02_STAGE03_RST_road_only_s_curve_service_lane.png", "Rest stop S curve service lane road module", "휴게소 S자 서비스 차로 길 모듈", "S-curve stage path", "S자 스테이지 길", road_notes_en("S-curve service-lane pavement, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("S자 휴게소 서비스 차로 포장, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop S-curve service lane road module", "S-curve service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-026", "STAGE03_RST", "03_rest_stop/03_STAGE03_RST_road_only_t_junction_service_lane.png", "Rest stop T junction service lane road module", "휴게소 T자 서비스 차로 길 모듈", "T-junction stage path", "T자 스테이지 길", road_notes_en("T-junction service-lane pavement, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("T자 휴게소 서비스 차로 포장, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop T-junction service lane road module", "T-junction service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-027", "STAGE03_RST", "03_rest_stop/04_STAGE03_RST_road_only_left_90_service_corner.png", "Rest stop left 90 service corner road module", "휴게소 왼쪽 90도 서비스 코너 길 모듈", "Left 90-degree stage path", "왼쪽 90도 스테이지 길", road_notes_en("left 90-degree service-lane corner, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("왼쪽 90도 휴게소 서비스 차로 코너, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop left 90-degree service corner road module", "left corner service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-028", "STAGE03_RST", "03_rest_stop/05_STAGE03_RST_road_only_right_90_service_corner.png", "Rest stop right 90 service corner road module", "휴게소 오른쪽 90도 서비스 코너 길 모듈", "Right 90-degree stage path", "오른쪽 90도 스테이지 길", road_notes_en("right 90-degree service-lane corner, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("오른쪽 90도 휴게소 서비스 차로 코너, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop right 90-degree service corner road module", "right corner service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-029", "STAGE03_RST", "03_rest_stop/06_STAGE03_RST_road_only_cross_service_intersection.png", "Rest stop cross service intersection road module", "휴게소 십자 서비스 교차로 길 모듈", "Cross intersection stage path", "십자 교차 스테이지 길", road_notes_en("cross service-lane intersection, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("십자 휴게소 서비스 차로 교차로, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop cross service intersection road module", "cross service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-030", "STAGE03_RST", "03_rest_stop/07_STAGE03_RST_road_only_roundabout_service_loop.png", "Rest stop roundabout service loop road module", "휴게소 원형 서비스 회전 길 모듈", "Roundabout stage path", "라운드어바웃 스테이지 길", road_notes_en("roundabout service-lane loop, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("원형 휴게소 서비스 차로 회전로, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop roundabout service loop road module", "roundabout service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-031", "STAGE03_RST", "03_rest_stop/08_STAGE03_RST_road_only_narrowing_service_lane.png", "Rest stop narrowing service lane road module", "휴게소 좁아지는 서비스 차로 길 모듈", "Narrowing stage path", "좁아지는 스테이지 길", road_notes_en("narrowing service-lane connector, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("좁아지는 휴게소 서비스 차로 연결로, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop narrowing service lane road module", "narrowing service pavement, curbs, parking seams, painted lines, and raised slab edges")),
        RoadAsset("RST-032", "STAGE03_RST", "03_rest_stop/09_STAGE03_RST_road_only_side_parking_bays.png", "Rest stop side parking bays road module", "휴게소 측면 주차 베이 길 모듈", "Side parking stage path", "측면 주차 스테이지 길", road_notes_en("service-lane road with side parking bays, soft concrete curbs, parking-lot seams, painted lines, and raised slab edges"), road_notes_kr("측면 주차 베이가 붙은 휴게소 서비스 차로, 낮은 콘크리트 커브, 주차장 이음선, 페인트 라인, 입체 슬래브 모서리"), brief("Rest stop side parking bays road module", "service pavement with side parking bays, curbs, parking seams, painted lines, and raised slab edges")),
    ],
    "STAGE04_CITY": [
        RoadAsset("CITY-030", "STAGE04_CITY", "04_city/01_STAGE04_CITY_road_only_straight_crosswalk_street.png", "City straight crosswalk street road module", "도시 직선 횡단보도 거리 길 모듈", "Straight stage path", "직선 스테이지 길", road_notes_en("city asphalt slab, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 아스팔트 슬래브, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City straight crosswalk street road module", "city asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-031", "STAGE04_CITY", "04_city/02_STAGE04_CITY_road_only_90_degree_corner_street.png", "City 90 degree corner street road module", "도시 90도 코너 거리 길 모듈", "90-degree stage path", "90도 스테이지 길", road_notes_en("city 90-degree asphalt corner, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 90도 아스팔트 코너, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City 90-degree corner street road module", "city corner asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-032", "STAGE04_CITY", "04_city/03_STAGE04_CITY_road_only_cross_intersection_street.png", "City cross intersection street road module", "도시 십자 교차로 거리 길 모듈", "Cross intersection stage path", "십자 교차 스테이지 길", road_notes_en("city cross intersection asphalt, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 십자 아스팔트 교차로, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City cross intersection street road module", "city cross intersection asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-033", "STAGE04_CITY", "04_city/04_STAGE04_CITY_road_only_t_junction_street.png", "City T junction street road module", "도시 T자 교차 거리 길 모듈", "T-junction stage path", "T자 스테이지 길", road_notes_en("city T-junction asphalt, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 T자 아스팔트 교차로, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City T-junction street road module", "city T-junction asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-034", "STAGE04_CITY", "04_city/05_STAGE04_CITY_road_only_s_curve_street.png", "City S curve street road module", "도시 S자 거리 길 모듈", "S-curve stage path", "S자 스테이지 길", road_notes_en("city S-curve asphalt, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 S자 아스팔트 도로, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City S-curve street road module", "city S-curve asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-035", "STAGE04_CITY", "04_city/06_STAGE04_CITY_road_only_y_split_street.png", "City Y split street road module", "도시 Y자 분기 거리 길 모듈", "Y-split stage path", "Y자 분기 스테이지 길", road_notes_en("city Y-split asphalt, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 Y자 아스팔트 분기, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City Y-split street road module", "city Y-split asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-036", "STAGE04_CITY", "04_city/07_STAGE04_CITY_road_only_narrowing_street.png", "City narrowing street road module", "도시 좁아지는 거리 길 모듈", "Narrowing stage path", "좁아지는 스테이지 길", road_notes_en("city narrowing asphalt connector, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 좁아지는 아스팔트 연결로, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City narrowing street road module", "city narrowing asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-037", "STAGE04_CITY", "04_city/08_STAGE04_CITY_road_only_roundabout_street.png", "City roundabout street road module", "도시 라운드어바웃 거리 길 모듈", "Roundabout stage path", "라운드어바웃 스테이지 길", road_notes_en("city roundabout asphalt, curbs, crosswalk paint, lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 라운드어바웃 아스팔트, 커브, 횡단보도 페인트, 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City roundabout street road module", "city roundabout asphalt, curbs, crosswalk paint, lane markings, manhole details, and sidewalk edges")),
        RoadAsset("CITY-038", "STAGE04_CITY", "04_city/09_STAGE04_CITY_road_only_bus_lane_street.png", "City bus lane street road module", "도시 버스전용차로 거리 길 모듈", "Bus lane stage path", "버스전용차로 스테이지 길", road_notes_en("city bus-lane asphalt, curbs, crosswalk paint, bus-lane markings, manhole details, and raised sidewalk edges"), road_notes_kr("도시 버스전용차로 아스팔트, 커브, 횡단보도 페인트, 버스 차선 표시, 맨홀 디테일, 입체 보도 모서리"), brief("City bus-lane street road module", "city bus-lane asphalt, curbs, crosswalk paint, bus-lane markings, manhole details, and sidewalk edges")),
    ],
    "STAGE05_GNG": [
        RoadAsset("GNG-031", "STAGE05_GNG", "05_gangnam/01_STAGE05_GNG_road_only_straight_luxury_boulevard.png", "Gangnam straight luxury boulevard road module", "강남 직선 럭셔리 대로 길 모듈", "Straight stage path", "직선 스테이지 길", road_notes_en("glossy dark boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 어두운 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam straight luxury boulevard road module", "glossy boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-032", "STAGE05_GNG", "05_gangnam/02_STAGE05_GNG_road_only_s_curve_luxury_boulevard.png", "Gangnam S curve luxury boulevard road module", "강남 S자 럭셔리 대로 길 모듈", "S-curve stage path", "S자 스테이지 길", road_notes_en("glossy S-curve boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 S자 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam S-curve luxury boulevard road module", "glossy S-curve asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-033", "STAGE05_GNG", "05_gangnam/03_STAGE05_GNG_road_only_y_split_luxury_boulevard.png", "Gangnam Y split luxury boulevard road module", "강남 Y자 럭셔리 대로 길 모듈", "Y-split stage path", "Y자 분기 스테이지 길", road_notes_en("glossy Y-split boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 Y자 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam Y-split luxury boulevard road module", "glossy Y-split asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-034", "STAGE05_GNG", "05_gangnam/04_STAGE05_GNG_road_only_left_90_luxury_boulevard.png", "Gangnam left 90 luxury boulevard road module", "강남 왼쪽 90도 럭셔리 대로 길 모듈", "Left 90-degree stage path", "왼쪽 90도 스테이지 길", road_notes_en("glossy left 90-degree boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 왼쪽 90도 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam left 90-degree luxury boulevard road module", "glossy left corner asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-035", "STAGE05_GNG", "05_gangnam/05_STAGE05_GNG_road_only_right_90_luxury_boulevard.png", "Gangnam right 90 luxury boulevard road module", "강남 오른쪽 90도 럭셔리 대로 길 모듈", "Right 90-degree stage path", "오른쪽 90도 스테이지 길", road_notes_en("glossy right 90-degree boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 오른쪽 90도 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam right 90-degree luxury boulevard road module", "glossy right corner asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-036", "STAGE05_GNG", "05_gangnam/06_STAGE05_GNG_road_only_t_junction_luxury_boulevard.png", "Gangnam T junction luxury boulevard road module", "강남 T자 럭셔리 대로 길 모듈", "T-junction stage path", "T자 스테이지 길", road_notes_en("glossy T-junction boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 T자 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam T-junction luxury boulevard road module", "glossy T-junction asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-037", "STAGE05_GNG", "05_gangnam/07_STAGE05_GNG_road_only_cross_luxury_plaza.png", "Gangnam cross luxury plaza road module", "강남 십자 럭셔리 플라자 길 모듈", "Cross plaza stage path", "십자 교차 스테이지 길", road_notes_en("glossy cross plaza boulevard asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 십자 플라자 대로 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam cross luxury plaza road module", "glossy cross plaza asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-038", "STAGE05_GNG", "05_gangnam/08_STAGE05_GNG_road_only_hairpin_valet_loop.png", "Gangnam hairpin valet loop road module", "강남 헤어핀 발렛 회차 길 모듈", "Hairpin valet stage path", "헤어핀 발렛 스테이지 길", road_notes_en("glossy hairpin valet loop asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 헤어핀 발렛 회차 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam hairpin valet loop road module", "glossy hairpin valet asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
        RoadAsset("GNG-039", "STAGE05_GNG", "05_gangnam/09_STAGE05_GNG_road_only_narrowing_luxury_entrance.png", "Gangnam narrowing luxury entrance road module", "강남 좁아지는 럭셔리 입구 길 모듈", "Narrowing entrance stage path", "좁아지는 입구 스테이지 길", road_notes_en("glossy narrowing luxury entrance asphalt, premium curb trim, reflective lane markings, tile insets, and raised slab edges"), road_notes_kr("광택 있는 좁아지는 럭셔리 입구 아스팔트, 고급 커브 트림, 반사 차선 표시, 타일 인셋, 입체 슬래브 모서리"), brief("Gangnam narrowing luxury entrance road module", "glossy narrowing entrance asphalt, premium curb trim, reflective lane markings, tile insets, and slab edges")),
    ],
}


def active_image_names() -> list[str]:
    return sorted(
        [path.name for path in IMAGE_DIR.glob("*.png") if path.is_file() and NUMBERED_IMAGE.match(path.name)],
        key=lambda name: int(name[:3]),
    )


def queue_asset_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    queue_ws = workbook.worksheets[-1]
    return {
        str(row[3]): str(row[1])
        for row in queue_ws.iter_rows(min_row=2, values_only=True)
        if row and row[1] and row[3]
    }


def build_entries(image_names: list[str], asset_by_filename: dict[str, str]) -> tuple[list[ActiveEntry], list[str]]:
    entries: list[ActiveEntry] = []
    archived_stage_roads: list[str] = []
    inserted_stages: set[str] = set()

    for filename in image_names:
        match = ACTIVE_IMAGE.match(filename)
        if not match:
            raise RuntimeError(f"Unexpected active image filename: {filename}")

        stage = match.group("stage")
        kind = match.group("kind")
        if stage in ROAD_ASSETS_BY_STAGE and kind == "ROAD":
            archived_stage_roads.append(filename)
            if stage not in inserted_stages:
                for road in ROAD_ASSETS_BY_STAGE[stage]:
                    entries.append(ActiveEntry(road.asset_id, road.name_en, road.rest_name, road=road))
                inserted_stages.add(stage)
            continue

        asset_id = asset_by_filename.get(filename)
        if not asset_id:
            raise RuntimeError(f"Workbook queue is missing active image filename: {filename}")
        entries.append(
            ActiveEntry(
                asset_id=asset_id,
                name_en=filename_name_en(filename),
                rest_name=filename[4:],
                old_filename=filename,
            )
        )

    missing_stages = sorted(set(ROAD_ASSETS_BY_STAGE) - inserted_stages)
    if missing_stages:
        raise RuntimeError(f"Did not find existing ROAD insertion points for: {missing_stages}")

    target_names = [entry.filename(index) for index, entry in enumerate(entries, start=1)]
    duplicates = sorted({name for name in target_names if target_names.count(name) > 1})
    if duplicates:
        raise RuntimeError(f"Duplicate target filenames: {duplicates[:8]}")

    return entries, archived_stage_roads


def filename_name_en(filename: str) -> str:
    match = ACTIVE_IMAGE.match(filename)
    if not match:
        return filename
    return match.group("name").replace("_", " ")


def copy_sheet_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)


def clear_and_write_rows(ws, rows: list[list[object]], template_row: int = 2) -> None:
    max_col = ws.max_column
    template_cells = [ws.cell(row=template_row, column=col) for col in range(1, max_col + 1)]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_index, values in enumerate(rows, start=2):
        for col_index in range(1, max_col + 1):
            cell = ws.cell(
                row=row_index,
                column=col_index,
                value=values[col_index - 1] if col_index <= len(values) else None,
            )
            copy_sheet_style(template_cells[col_index - 1], cell)


def format_audit_reference(match: re.Match[str], sequence_by_asset: dict[str, int]) -> str:
    asset_id = match.group(1)
    sequence = sequence_by_asset.get(asset_id)
    if sequence is None:
        return asset_id
    return f"{sequence:03d}_{asset_id}"


def sync_audit_sheet(workbook, sequence_by_asset: dict[str, int]) -> None:
    if len(workbook.worksheets) < 4:
        return
    audit_ws = workbook.worksheets[3]
    if audit_ws.max_column < 4:
        return
    for row_index in range(2, audit_ws.max_row + 1):
        cell = audit_ws.cell(row=row_index, column=4)
        if isinstance(cell.value, str):
            cell.value = AUDIT_ASSET_REF.sub(
                lambda match: format_audit_reference(match, sequence_by_asset),
                cell.value,
            )


def road_row(road: RoadAsset, korean: bool) -> list[object]:
    region_en = {
        "STAGE01_NRY": "Noryangjin",
        "STAGE02_HWY": "Highway",
        "STAGE03_RST": "Rest Stop",
        "STAGE04_CITY": "City",
        "STAGE05_GNG": "Gangnam",
    }[road.stage_code]
    region_kr = {
        "STAGE01_NRY": "노량진",
        "STAGE02_HWY": "고속도로",
        "STAGE03_RST": "휴게소",
        "STAGE04_CITY": "도시",
        "STAGE05_GNG": "강남",
    }[road.stage_code]

    return [
        road.asset_id,
        region_kr if korean else region_en,
        road.name_kr,
        road.name_en,
        "도로 모듈" if korean else "Road Module",
        road.role_kr if korean else road.role_en,
        "P0",
        "중간" if korean else "Medium",
        "Image to 3D",
        "스테이지별 road-only 길 모듈 변형" if korean else "Stage-specific road-only module variants",
        "전체 난이도" if korean else "All difficulties",
        road.visual_notes_kr if korean else road.visual_notes_en,
        road.brief,
    ]


def update_workbook(path: Path, korean: bool, entries: list[ActiveEntry]) -> None:
    workbook = load_workbook(path)
    asset_ws = workbook.worksheets[1]
    queue_ws = workbook.worksheets[-1]

    row_by_asset = {
        str(row[0]): list(row)
        for row in asset_ws.iter_rows(min_row=2, max_col=asset_ws.max_column, values_only=True)
        if row and row[0]
    }

    asset_rows: list[list[object]] = []
    queue_rows: list[list[object]] = []
    sequence_by_asset: dict[str, int] = {}
    status_value = "생성됨" if korean else "Generated"

    for sequence, entry in enumerate(entries, start=1):
        if entry.road is not None:
            asset_row = road_row(entry.road, korean)
        else:
            asset_row = row_by_asset.get(entry.asset_id)
            if asset_row is None:
                raise RuntimeError(f"{path.name}: missing workbook asset row for {entry.asset_id}")
        filename = entry.filename(sequence)
        asset_rows.append(asset_row)
        queue_rows.append([sequence, entry.asset_id, asset_row[3], filename, status_value])
        sequence_by_asset[entry.asset_id] = sequence

    clear_and_write_rows(asset_ws, asset_rows)
    clear_and_write_rows(queue_ws, queue_rows)
    sync_audit_sheet(workbook, sequence_by_asset)
    workbook.save(path)


def write_korean_jsonl(entries: list[ActiveEntry]) -> None:
    workbook = load_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", read_only=True, data_only=True)
    asset_ws = workbook.worksheets[1]
    row_by_asset = {
        str(row[0]): row
        for row in asset_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    }

    output = DESIGN_DIR / "meshy_image_prompts_kr.jsonl"
    with output.open("w", encoding="utf-8", newline="\n") as handle:
        for sequence, entry in enumerate(entries, start=1):
            row = row_by_asset[entry.asset_id]
            payload = {
                "sequence": sequence,
                "asset_id": row[0],
                "region": row[1],
                "name_kr": row[2],
                "name_en": row[3],
                "type": row[4],
                "role": row[5],
                "priority": row[6],
                "meshy_fit": row[7],
                "recommended_input": row[8],
                "visual_notes": row[11],
                "brief": row[12],
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def validate_sources(entries: list[ActiveEntry]) -> None:
    missing_sources = sorted(
        str(entry.road.source_path.relative_to(ROOT))
        for entry in entries
        if entry.road is not None and not entry.road.source_path.exists()
    )
    if missing_sources:
        raise RuntimeError(f"Missing RnD road source image(s): {missing_sources}")

    missing_kept = sorted(
        str((IMAGE_DIR / entry.old_filename).relative_to(ROOT))
        for entry in entries
        if entry.old_filename and not (IMAGE_DIR / entry.old_filename).exists()
    )
    if missing_kept:
        raise RuntimeError(f"Missing active image(s) to keep: {missing_kept[:8]}")


def apply_image_changes(entries: list[ActiveEntry], archived_stage_roads: list[str]) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_dir = IMAGE_DIR / "old" / f"stage_roads_replaced_{timestamp}"
    archive_dir.mkdir(parents=True, exist_ok=False)

    temp_paths: list[tuple[Path, Path]] = []
    for entry in entries:
        if not entry.old_filename:
            continue
        source = IMAGE_DIR / entry.old_filename
        temp = IMAGE_DIR / f".__road_promote_tmp__{entry.old_filename}"
        if temp.exists():
            raise RuntimeError(f"Temporary path already exists: {temp}")
        temp_paths.append((temp, source))

    target_paths = [IMAGE_DIR / entry.filename(index) for index, entry in enumerate(entries, start=1)]
    existing_unexpected_targets = [
        path
        for path in target_paths
        if path.exists() and path.name not in {entry.old_filename for entry in entries if entry.old_filename}
    ]
    if existing_unexpected_targets:
        raise RuntimeError(f"Target path(s) already exist unexpectedly: {existing_unexpected_targets[:8]}")

    for temp, source in temp_paths:
        source.replace(temp)

    for filename in archived_stage_roads:
        source = IMAGE_DIR / filename
        if source.exists():
            source.replace(archive_dir / filename)

    for sequence, entry in enumerate(entries, start=1):
        target = IMAGE_DIR / entry.filename(sequence)
        if entry.old_filename:
            temp = IMAGE_DIR / f".__road_promote_tmp__{entry.old_filename}"
            temp.replace(target)
        elif entry.road is not None:
            shutil.copy2(entry.road.source_path, target)
        else:
            raise RuntimeError(f"Entry has neither old file nor road source: {entry}")

    return archive_dir


def summarize(entries: list[ActiveEntry], archived_stage_roads: list[str]) -> None:
    stage_counts = {
        stage: sum(1 for entry in entries if entry.road and entry.road.stage_code == stage)
        for stage in ROAD_ASSETS_BY_STAGE
    }
    print(f"active image target count: {len(entries)}")
    print(f"old stage ROAD images removed from active root: {len(archived_stage_roads)}")
    for stage, count in stage_counts.items():
        print(f"{stage}: {count} promoted RnD road-only images")
    print(f"new sequence range: 001-{len(entries):03d}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Promote approved RnD road-only images into the active Meshy image queue."
    )
    parser.add_argument("--apply", action="store_true", help="Apply image, workbook, and JSONL changes.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    image_names = active_image_names()
    if not image_names:
        raise RuntimeError("No active numbered PNG images found.")

    asset_by_filename = queue_asset_map(DESIGN_DIR / "tralalero_meshy_asset_plan.xlsx")
    entries, archived_stage_roads = build_entries(image_names, asset_by_filename)
    validate_sources(entries)
    summarize(entries, archived_stage_roads)

    if not args.apply:
        print("dry run only; pass --apply to write files")
        return

    archive_dir = apply_image_changes(entries, archived_stage_roads)
    for workbook, korean in WORKBOOKS:
        update_workbook(workbook, korean, entries)
        print(f"updated {workbook.relative_to(ROOT)}")
    write_korean_jsonl(entries)
    print(f"updated {(DESIGN_DIR / 'meshy_image_prompts_kr.jsonl').relative_to(ROOT)}")
    print(f"archived old active stage ROAD images under {archive_dir.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
