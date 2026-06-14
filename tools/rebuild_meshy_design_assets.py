from __future__ import annotations

import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DESIGN_DIR = ROOT / "docs" / "design"
IMAGE_DIR = ROOT / "output" / "meshy_images"
ASSET_ID_PATTERN = re.compile(r"(?:^|_)([A-Z]+-\d{3})(?:_|\.png$)")

STAGE_CODES = {
    "Noryangjin": "STAGE01_NRY",
    "노량진": "STAGE01_NRY",
    "Highway": "STAGE02_HWY",
    "고속도로": "STAGE02_HWY",
    "Rest Stop": "STAGE03_RST",
    "휴게소": "STAGE03_RST",
    "City": "STAGE04_CITY",
    "도시": "STAGE04_CITY",
    "Gangnam": "STAGE05_GNG",
    "강남": "STAGE05_GNG",
    "Common": "COMMON",
    "공통": "COMMON",
}

PREFIX_STAGE_CODES = {
    "NRY": "STAGE01_NRY",
    "HWY": "STAGE02_HWY",
    "RST": "STAGE03_RST",
    "CITY": "STAGE04_CITY",
    "GNG": "STAGE05_GNG",
}

ROAD_STAGE_CODES = {
    "ROAD-001": "STAGE01_NRY",
    "ROAD-002": "STAGE01_NRY",
    "ROAD-003": "STAGE02_HWY",
    "ROAD-004": "STAGE04_CITY",
    "ROAD-005": "STAGE01_NRY",
    "ROAD-006": "STAGE02_HWY",
    "ROAD-007": "COMMON",
    "ROAD-008": "COMMON",
}

STAGE_ROAD_TILE_ASSET_IDS = {
    "NRY-015",
    "NRY-028",
    "NRY-035",
    "RST-007",
    "RST-021",
    "CITY-009",
    "CITY-010",
    "CITY-020",
    "GNG-013",
    "GNG-020",
}

STAGE_PROP_ASSET_IDS = {
    "CITY-004",
}

EXISTING_ROAD_NOTE_OVERRIDES = {
    "ROAD-003": (
        "Clean dark highway asphalt floor tile, subtle surface wear only",
        "Single highway asphalt floor tile, plain clean dark slab, subtle surface wear, white background.",
    ),
    "ROAD-006": (
        "Clean dark highway asphalt strip tile, subtle surface wear only",
        "Single highway asphalt strip tile, plain clean dark slab, subtle surface wear, white background.",
    ),
    "ROAD-004": (
        "Clean muted city pavement floor tile, subtle surface wear only",
        "Single city pavement floor tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "RST-007": (
        "Clean muted rest-stop pavement slab, subtle surface wear only",
        "Single rest-stop pavement floor tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "RST-021": (
        "Clean muted rest-stop service pavement slab, subtle surface wear only",
        "Single rest-stop service pavement floor tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "CITY-009": (
        "Clean muted city asphalt floor tile, subtle surface wear only",
        "Single city asphalt floor tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "CITY-010": (
        "Clean muted city asphalt detail tile, subtle surface wear only",
        "Single city asphalt detail tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "CITY-020": (
        "Clean muted city asphalt service tile, subtle surface wear only",
        "Single city asphalt service tile, plain clean muted slab, subtle surface wear, white background.",
    ),
    "GNG-013": (
        "Clean glossy dark Gangnam boulevard floor tile, subtle surface wear only",
        "Single Gangnam boulevard floor tile, plain clean glossy dark slab, subtle surface wear, white background.",
    ),
    "GNG-020": (
        "Clean glossy dark Gangnam boulevard entrance tile, subtle surface wear only",
        "Single Gangnam boulevard entrance tile, plain clean glossy dark slab, subtle surface wear, white background.",
    ),
}

EXISTING_ROAD_NAME_OVERRIDES = {
    "ROAD-003": (
        "\uace0\uc18d\ub3c4\ub85c \uae30\ubcf8 \uc544\uc2a4\ud314\ud2b8 \ubc14\ub2e5 \ud0c0\uc77c",
        "Highway plain asphalt floor tile",
    ),
    "ROAD-006": (
        "\uace0\uc18d\ub3c4\ub85c \uae30\ubcf8 \uc544\uc2a4\ud314\ud2b8 \uc2a4\ud2b8\ub9bd \ud0c0\uc77c",
        "Highway plain asphalt strip tile",
    ),
    "ROAD-004": (
        "\ub3c4\uc2dc \uae30\ubcf8 \ud3ec\uc7a5 \ubc14\ub2e5 \ud0c0\uc77c",
        "City plain pavement floor tile",
    ),
    "RST-007": (
        "\ud734\uac8c\uc18c \uae30\ubcf8 \ud3ec\uc7a5 \ubc14\ub2e5 \ud0c0\uc77c",
        "Rest stop plain pavement floor tile",
    ),
    "RST-021": (
        "\ud734\uac8c\uc18c \uae30\ubcf8 \uc11c\ube44\uc2a4\ub85c \ubc14\ub2e5 \ud0c0\uc77c",
        "Rest stop plain service pavement tile",
    ),
    "CITY-009": (
        "\ub3c4\uc2dc \uae30\ubcf8 \uc544\uc2a4\ud314\ud2b8 \ubc14\ub2e5 \ud0c0\uc77c",
        "City plain asphalt floor tile",
    ),
    "CITY-010": (
        "\ub3c4\uc2dc \uae30\ubcf8 \uc544\uc2a4\ud314\ud2b8 \ub514\ud14c\uc77c \ud0c0\uc77c",
        "City plain asphalt detail tile",
    ),
    "CITY-020": (
        "\ub3c4\uc2dc \uae30\ubcf8 \uc544\uc2a4\ud314\ud2b8 \uc11c\ube44\uc2a4 \ud0c0\uc77c",
        "City plain asphalt service tile",
    ),
    "GNG-013": (
        "\uac15\ub0a8 \uae30\ubcf8 \uc5b4\ub450\uc6b4 \ub300\ub85c \ubc14\ub2e5 \ud0c0\uc77c",
        "Gangnam plain dark boulevard floor tile",
    ),
    "GNG-020": (
        "\uac15\ub0a8 \uae30\ubcf8 \uc5b4\ub450\uc6b4 \ub300\ub85c \uc785\uad6c \ud0c0\uc77c",
        "Gangnam plain dark boulevard entrance tile",
    ),
}

STAGE_SORT_ORDER = {
    "STAGE01_NRY": 0,
    "STAGE02_HWY": 1,
    "STAGE03_RST": 2,
    "STAGE04_CITY": 3,
    "STAGE05_GNG": 4,
    "COMMON": 5,
}


ADDITIONAL_ASSETS = [
    {
        "asset_id": "NRY-036",
        "region_en": "Noryangjin",
        "region_kr": "노량진",
        "name_kr": "수산시장 나무 X 바리케이드",
        "name_en": "Fish-market wooden X barricade",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Lane blocker",
        "use_kr": "차선 차단 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Noryangjin wood blocker color variants",
        "reuse_kr": "노량진 목재 차단물 색상 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Chunky wood frame, red-white X board, wet market scuffs",
        "brief": "Single fish-market wooden X barricade for a stylized mobile runner, wet Noryangjin market style, chunky readable obstacle, white background, 3/4 view.",
    },
    {
        "asset_id": "NRY-037",
        "region_en": "Noryangjin",
        "region_kr": "노량진",
        "name_kr": "항구 차선 신호 게이트",
        "name_en": "Harbor lane signal gantry",
        "category_en": "Gameplay Module",
        "category_kr": "게임플레이 모듈",
        "use_en": "Stage direction cue",
        "use_kr": "구간 진행 방향 신호",
        "priority": "P1",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Noryangjin bridge and stage-entry variants",
        "reuse_kr": "노량진 다리/진입 게이트 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Thick blue metal posts, green arrow panels, small traffic light",
        "brief": "Single harbor lane signal gantry with thick posts, green arrow panels and a compact traffic light, stylized 3D mobile runner prop, white background.",
    },
    {
        "asset_id": "HWY-023",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "이동식 LED 화살표 트레일러",
        "name_en": "Portable LED arrow trailer",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Construction lane cue",
        "use_kr": "공사 차선 유도 장애물",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Arrow-board variants for construction clusters",
        "reuse_kr": "공사 구간 화살표 보드 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Black LED board, amber arrows, chunky trailer base",
        "brief": "Single portable LED arrow board trailer, highway construction warning prop, amber arrow lights, chunky trailer base, stylized 3D, white background.",
    },
    {
        "asset_id": "HWY-024",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "초록 탄성 차선 분리봉 라인",
        "name_en": "Green flexible lane bollard row",
        "category_en": "Boundary",
        "category_kr": "경계물",
        "use_en": "Lane separator",
        "use_kr": "차선 분리 경계",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Short and long separator rows",
        "reuse_kr": "짧은/긴 분리봉 줄 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Several thick green posts on a low curb base",
        "brief": "Single row of flexible green highway lane bollards on a low curb base, stylized 3D mobile runner boundary prop, white background.",
    },
    {
        "asset_id": "HWY-025",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "톨게이트 차단봉 모듈",
        "name_en": "Toll barrier arm module",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Gate blocker",
        "use_kr": "요금소 차단 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Open, closed, and broken arm variants",
        "reuse_kr": "열림/닫힘/파손 차단봉 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Red-white arm, heavy yellow-black base, readable clearance",
        "brief": "Single toll barrier arm module with a red-white striped arm and chunky yellow-black base, stylized 3D highway runner obstacle, white background.",
    },
    {
        "asset_id": "HWY-026",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "톨게이트 차선 신호등 모듈",
        "name_en": "Toll lane signal light module",
        "category_en": "Gameplay Module",
        "category_kr": "게임플레이 모듈",
        "use_en": "Lane state read",
        "use_kr": "통과 가능 차선 표시",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Red X and green arrow variants",
        "reuse_kr": "빨간 X/초록 화살표 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Chunky overhead signal boxes, bright red X and green arrows",
        "brief": "Single toll lane signal module with chunky overhead boxes showing red X and green arrow lights, stylized 3D highway prop, white background.",
    },
    {
        "asset_id": "HWY-027",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "요금소 결제 단말기 섬",
        "name_en": "Toll payment terminal island",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Toll booth detail",
        "use_kr": "요금소 디테일",
        "priority": "P1",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Toll plaza side-detail variants",
        "reuse_kr": "요금소 측면 디테일 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Payment post on island, card slot, green/red indicator, thick curb",
        "brief": "Single highway toll payment terminal on a chunky concrete island with card slot and small indicator lights, stylized 3D prop, white background.",
    },
    {
        "asset_id": "RST-019",
        "region_en": "Rest Stop",
        "region_kr": "휴게소",
        "name_kr": "분리수거 쓰레기통 3종",
        "name_en": "Recycling bin trio",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Side decoration",
        "use_kr": "차선 옆 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Color-coded bin variants",
        "reuse_kr": "색상별 쓰레기통 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Blue, green, orange bins, large top openings, no small text",
        "brief": "Single grouped trio of color-coded recycling bins for a Korean rest stop, chunky stylized 3D mobile runner prop, white background.",
    },
    {
        "asset_id": "RST-020",
        "region_en": "Rest Stop",
        "region_kr": "휴게소",
        "name_kr": "카페 메뉴 입간판",
        "name_en": "Cafe menu sandwich board",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Shop cue",
        "use_kr": "매장 유도 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Coffee, snack, and restroom icon variants",
        "reuse_kr": "커피/간식/화장실 아이콘 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "A-frame board, big cup icon, no readable text",
        "brief": "Single cafe menu A-frame sandwich board with a large simple cup icon, stylized 3D rest-stop prop, no readable text, white background.",
    },
    {
        "asset_id": "RST-021",
        "region_en": "Rest Stop",
        "region_kr": "휴게소",
        "name_kr": "휴게소 기본 서비스로 바닥 타일",
        "name_en": "Rest stop plain service pavement tile",
        "category_en": "Road Tile",
        "category_kr": "바닥 타일",
        "use_en": "Service-road floor read",
        "use_kr": "서비스로 바닥 판독",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Rest-stop pavement variants",
        "reuse_kr": "휴게소 포장 바닥 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Clean muted rest-stop service pavement slab, subtle surface wear only",
        "brief": "Single rest-stop service pavement floor tile, plain clean muted slab, subtle surface wear, white background.",
    },
    {
        "asset_id": "CITY-024",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "보행자 신호 버튼 기둥",
        "name_en": "Pedestrian signal button pole",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Crosswalk detail",
        "use_kr": "횡단보도 디테일",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Crosswalk signal detail variants",
        "reuse_kr": "횡단보도 신호 디테일 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Small pole, green walking icon, chunky button box",
        "brief": "Single pedestrian signal button pole with a green walking icon and chunky push-button box, stylized 3D city prop, white background.",
    },
    {
        "asset_id": "CITY-025",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "공사장 굴착기 팔 소품",
        "name_en": "Construction excavator arm prop",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Construction side hazard",
        "use_kr": "공사 구역 위험물",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Behind-fence and lane-blocker variants",
        "reuse_kr": "펜스 뒤 배치/차선 차단 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Orange excavator boom and bucket, thick simplified joints",
        "brief": "Single orange construction excavator arm with bucket, simplified chunky joints for a stylized city runner obstacle, white background.",
    },
    {
        "asset_id": "CITY-026",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "세로 거리 배너 기둥 세트",
        "name_en": "Vertical street banner pole set",
        "category_en": "Decoration",
        "category_kr": "장식물",
        "use_en": "Street rhythm decoration",
        "use_kr": "거리 리듬 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Color banner variants across city and Gangnam",
        "reuse_kr": "도시/강남 색상 배너 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Thick pole, stacked bright vertical panels, no text",
        "brief": "Single street pole with stacked vertical color banner panels, stylized 3D urban decoration, no text, white background.",
    },
    {
        "asset_id": "GNG-023",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "고급 세로 LED 배너 기둥",
        "name_en": "Luxury vertical LED banner pillar",
        "category_en": "Decoration",
        "category_kr": "장식물",
        "use_en": "Luxury district branding",
        "use_kr": "고급 상권 브랜딩 장식",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Purple-blue gradient panel variants",
        "reuse_kr": "보라/파랑 패널 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Tall black-gold frame, glowing gradient screen, no text",
        "brief": "Single luxury vertical LED banner pillar with black-gold frame and glowing purple-blue gradient screen, stylized 3D Gangnam prop, no text, white background.",
    },
    {
        "asset_id": "GNG-024",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "벨벳 로프 스탠션 세트",
        "name_en": "Velvet rope stanchion set",
        "category_en": "Boundary",
        "category_kr": "경계물",
        "use_en": "Luxury entry boundary",
        "use_kr": "고급 입구 경계",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Queue, red carpet, and reward entry variants",
        "reuse_kr": "대기줄/레드카펫/보상 입구 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Gold posts, thick red velvet ropes, readable spacing",
        "brief": "Single velvet rope stanchion set with gold posts and thick red ropes, luxury Gangnam entry boundary prop, stylized 3D, white background.",
    },
    {
        "asset_id": "GNG-025",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "주얼리 유리 진열장",
        "name_en": "Jewelry glass display case",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Premium display obstacle",
        "use_kr": "고급 진열 장애물",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Jewelry, watch, and handbag display variants",
        "reuse_kr": "주얼리/시계/가방 진열 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Black base, gold trim, simplified glass box, necklace bust",
        "brief": "Single luxury jewelry glass display case with black base, gold trim, and a simple necklace bust inside, stylized 3D, white background.",
    },
    {
        "asset_id": "GNG-026",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "명품 핸드백 진열대",
        "name_en": "Luxury handbag display plinth",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Premium display obstacle",
        "use_kr": "고급 진열 장애물",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Handbag and shoe plinth variants",
        "reuse_kr": "핸드백/신발 진열대 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Marble plinth, one handbag, clean silhouette",
        "brief": "Single luxury handbag displayed on a black marble plinth with gold trim, stylized 3D Gangnam runner prop, white background.",
    },
    {
        "asset_id": "GNG-027",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "다이아몬드 로고 조형물",
        "name_en": "Diamond logo sculpture sign",
        "category_en": "Landmark",
        "category_kr": "랜드마크",
        "use_en": "Department-store landmark",
        "use_kr": "백화점 랜드마크",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Final entrance and side-plaza variants",
        "reuse_kr": "최종 입구/광장 장식 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Black-gold diamond mark, thick base, no readable text",
        "brief": "Single black-gold diamond logo sculpture sign on a heavy pedestal, luxury department-store landmark prop, stylized 3D, no text, white background.",
    },
    {
        "asset_id": "GNG-028",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "부티크 벽부등 모듈",
        "name_en": "Boutique wall sconce light module",
        "category_en": "Decoration",
        "category_kr": "장식물",
        "use_en": "Luxury lighting detail",
        "use_kr": "고급 조명 디테일",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Warm storefront lighting variants",
        "reuse_kr": "따뜻한 매장 조명 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Gold fixture, warm lamp, thick back plate",
        "brief": "Single boutique wall sconce light module with gold fixture and warm lamp on a thick back plate, stylized 3D Gangnam storefront prop, white background.",
    },
    {
        "asset_id": "HWY-028",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "파란 고속버스 장애물",
        "name_en": "Blue highway bus obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Large traffic obstacle",
        "use_kr": "대형 교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Color and route variants for highway and city stages",
        "reuse_kr": "고속도로/도시 버스 색상 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Blue city-bus body, large windshield, chunky wheels",
        "brief": "Single blue highway bus obstacle for a stylized mobile runner, chunky readable vehicle shape, white background, 3/4 view.",
    },
    {
        "asset_id": "HWY-029",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "빨간 소형 승용차 장애물",
        "name_en": "Red compact passenger car obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Traffic obstacle",
        "use_kr": "교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Color variants from the passenger car base",
        "reuse_kr": "승용차 기본 모델 색상 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Red/orange compact hatchback, simple windshield, thick wheels",
        "brief": "Single red compact passenger car obstacle for a highway runner stage, stylized 3D, chunky readable silhouette, white background.",
    },
    {
        "asset_id": "HWY-030",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "초록 소형 승용차 장애물",
        "name_en": "Green compact passenger car obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Traffic obstacle",
        "use_kr": "교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Color variants from the passenger car base",
        "reuse_kr": "승용차 기본 모델 색상 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Green compact hatchback, simple windshield, thick wheels",
        "brief": "Single green compact passenger car obstacle for a highway runner stage, stylized 3D, chunky readable silhouette, white background.",
    },
    {
        "asset_id": "HWY-031",
        "region_en": "Highway",
        "region_kr": "고속도로",
        "name_kr": "노란 박스 트럭 장애물",
        "name_en": "Yellow box truck obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Large traffic obstacle",
        "use_kr": "대형 교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Truck color and delivery variants",
        "reuse_kr": "트럭 색상/배송차 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Yellow cargo box, cab, chunky rear doors, thick wheels",
        "brief": "Single yellow box truck obstacle for a Korean highway runner stage, stylized 3D, chunky readable vehicle shape, white background.",
    },
    {
        "asset_id": "RST-022",
        "region_en": "Rest Stop",
        "region_kr": "휴게소",
        "name_kr": "흰색 주차 SUV 소품",
        "name_en": "White parked SUV prop",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Parking lot decoration",
        "use_kr": "주차장 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Parked car and SUV variants",
        "reuse_kr": "주차 차량/SUV 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "White SUV, side-parked stance, rounded body",
        "brief": "Single white parked SUV prop for a Korean rest stop parking lot, stylized 3D, chunky readable vehicle, white background.",
    },
    {
        "asset_id": "RST-023",
        "region_en": "Rest Stop",
        "region_kr": "휴게소",
        "name_kr": "휴게소 셔틀버스 소품",
        "name_en": "Rest-stop shuttle bus prop",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Parking lot traffic decoration",
        "use_kr": "주차장 교통 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Rest stop and city bus variants",
        "reuse_kr": "휴게소/도시 버스 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Small blue shuttle bus, large windows, rounded body",
        "brief": "Single small blue rest-stop shuttle bus prop, stylized 3D mobile runner vehicle, chunky readable shape, white background.",
    },
    {
        "asset_id": "CITY-027",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "도심 파란 버스 장애물",
        "name_en": "City blue bus obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Large urban traffic obstacle",
        "use_kr": "대형 도심 교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "City bus obstacle and parked variant",
        "reuse_kr": "도심 버스 장애물/정차 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Blue urban bus, front read, wide windshield",
        "brief": "Single blue city bus obstacle for an urban runner stage, stylized 3D, chunky readable vehicle silhouette, white background.",
    },
    {
        "asset_id": "CITY-028",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "노란 도심 택시 장애물",
        "name_en": "Yellow city taxi obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Traffic obstacle",
        "use_kr": "교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Taxi color variants",
        "reuse_kr": "택시 색상 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Yellow taxi, roof light, compact sedan silhouette",
        "brief": "Single yellow city taxi obstacle with roof light, stylized 3D mobile runner vehicle, chunky readable silhouette, white background.",
    },
    {
        "asset_id": "CITY-029",
        "region_en": "City",
        "region_kr": "도시",
        "name_kr": "배달 라이더 스쿠터 장애물",
        "name_en": "Delivery rider scooter obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Moving urban obstacle",
        "use_kr": "도심 이동 장애물",
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "중간",
        "method": "Image to 3D",
        "reuse_en": "Scooter-only and rider variants",
        "reuse_kr": "스쿠터 단독/라이더 포함 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Green helmet rider, delivery box, chunky scooter",
        "brief": "Single delivery rider scooter obstacle with green helmet and rear delivery box, stylized 3D city runner vehicle, white background.",
    },
    {
        "asset_id": "GNG-029",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "강남 검은 고급 세단 장애물",
        "name_en": "Gangnam black luxury sedan obstacle",
        "category_en": "Obstacle",
        "category_kr": "장애물",
        "use_en": "Luxury traffic obstacle",
        "use_kr": "고급 교통 장애물",
        "priority": "P0",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Valet and parked sedan variants",
        "reuse_kr": "발렛/주차 세단 변형",
        "difficulty_en": "Normal and above",
        "difficulty_kr": "노말 이상",
        "visual_notes": "Black sedan, glossy finish, gold reflections, low wide shape",
        "brief": "Single black luxury sedan obstacle for a Gangnam runner stage, stylized 3D, glossy chunky readable vehicle, white background.",
    },
    {
        "asset_id": "GNG-030",
        "region_en": "Gangnam",
        "region_kr": "강남",
        "name_kr": "강남 발렛 차량 소품",
        "name_en": "Gangnam valet car prop",
        "category_en": "Prop",
        "category_kr": "소품",
        "use_en": "Luxury entrance decoration",
        "use_kr": "고급 입구 장식",
        "priority": "P1",
        "meshy_en": "High",
        "meshy_kr": "높음",
        "method": "Image to 3D",
        "reuse_en": "Side-parked valet car variants",
        "reuse_kr": "측면 주차 발렛 차량 변형",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "전체 난이도",
        "visual_notes": "Black premium car with valet sign base, side-parked stance",
        "brief": "Single black Gangnam valet car prop with a small premium valet base cue, stylized 3D luxury vehicle, white background.",
    },
]


def road_module_asset(
    asset_id: str,
    region_en: str,
    region_kr: str,
    name_en: str,
    name_kr: str,
    use_en: str,
    use_kr: str,
    visual_notes: str,
    brief: str,
) -> dict[str, str]:
    return {
        "asset_id": asset_id,
        "region_en": region_en,
        "region_kr": region_kr,
        "name_kr": name_kr,
        "name_en": name_en,
        "category_en": "Road Module",
        "category_kr": "\ub3c4\ub85c \ubaa8\ub4c8",
        "use_en": use_en,
        "use_kr": use_kr,
        "priority": "P0",
        "meshy_en": "Medium",
        "meshy_kr": "\uc911\uac04",
        "method": "Image to 3D",
        "reuse_en": "Stage-specific path module variants",
        "reuse_kr": "\uc2a4\ud14c\uc774\uc9c0\ubcc4 \uae38 \ubaa8\ub4c8 \ubcc0\ud615",
        "difficulty_en": "All difficulties",
        "difficulty_kr": "\uc804\uccb4 \ub09c\uc774\ub3c4",
        "visual_notes": visual_notes,
        "brief": brief,
    }


ROAD_MODULE_ASSETS = [
    road_module_asset(
        "NRY-038",
        "Noryangjin",
        "\ub178\ub7c9\uc9c4",
        "Noryangjin straight dock path module",
        "\ub178\ub7c9\uc9c4 \uc9c1\uc120 \ubd80\ub450 \uae38 \ubaa8\ub4c8",
        "Straight stage path",
        "\uc9c1\uc120 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Wet fish-market concrete, dock planks, puddles, blue drainage edge",
        "Single straight Noryangjin wet market path module, dock planks and concrete lane, readable runner road tile, white background, 3/4 view.",
    ),
    road_module_asset(
        "NRY-039",
        "Noryangjin",
        "\ub178\ub7c9\uc9c4",
        "Noryangjin left bend dock path module",
        "\ub178\ub7c9\uc9c4 \uc67c\ucabd \ucee4\ube0c \ubd80\ub450 \uae38 \ubaa8\ub4c8",
        "Left bend stage path",
        "\uc67c\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Left-curving wet market lane with wooden pier trim and fish-market puddles",
        "Single left curve Noryangjin path module for a mobile runner, wet concrete and dock wood edges, white background, 3/4 view.",
    ),
    road_module_asset(
        "NRY-040",
        "Noryangjin",
        "\ub178\ub7c9\uc9c4",
        "Noryangjin right bend dock path module",
        "\ub178\ub7c9\uc9c4 \uc624\ub978\ucabd \ucee4\ube0c \ubd80\ub450 \uae38 \ubaa8\ub4c8",
        "Right bend stage path",
        "\uc624\ub978\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Right-curving wet market lane with blue plastic crate marks and puddles",
        "Single right curve Noryangjin path module for a mobile runner, wet market concrete, blue crate marks, white background, 3/4 view.",
    ),
    road_module_asset(
        "NRY-041",
        "Noryangjin",
        "\ub178\ub7c9\uc9c4",
        "Noryangjin narrow fish market lane module",
        "\ub178\ub7c9\uc9c4 \uc881\uc544\uc9c0\ub294 \uc218\uc0b0\uc2dc\uc7a5 \uae38 \ubaa8\ub4c8",
        "Narrowing stage path",
        "\uc881\uc544\uc9c0\ub294 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Narrow market lane, wet concrete, crate edges closing the path",
        "Single narrowing Noryangjin fish-market lane module, clear runner bottleneck shape, wet floor, white background, 3/4 view.",
    ),
    road_module_asset(
        "NRY-042",
        "Noryangjin",
        "\ub178\ub7c9\uc9c4",
        "Noryangjin split wet market lane module",
        "\ub178\ub7c9\uc9c4 \ubd84\uae30 \uc218\uc0b0\uc2dc\uc7a5 \uae38 \ubaa8\ub4c8",
        "Split stage path",
        "\ubd84\uae30 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Y-split wet market path with dock planks and seafood stall edge cues",
        "Single split-lane Noryangjin path module, wet fish market floor, readable Y branch, white background, 3/4 view.",
    ),
    road_module_asset(
        "HWY-032",
        "Highway",
        "\uace0\uc18d\ub3c4\ub85c",
        "Highway straight plain asphalt module",
        "\uace0\uc18d\ub3c4\ub85c \uc9c1\uc120 \uc544\uc2a4\ud314\ud2b8 \uae38 \ubaa8\ub4c8",
        "Straight stage path",
        "\uc9c1\uc120 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean dark asphalt slab, beveled road body, subtle surface wear only",
        "Single straight highway asphalt path module, plain clean dark road slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "HWY-033",
        "Highway",
        "\uace0\uc18d\ub3c4\ub85c",
        "Highway left bend plain asphalt module",
        "\uace0\uc18d\ub3c4\ub85c \uc67c\ucabd \ucee4\ube0c \uc544\uc2a4\ud314\ud2b8 \ubaa8\ub4c8",
        "Left bend stage path",
        "\uc67c\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean left-curving asphalt slab with subtle surface wear only",
        "Single left curve highway path module, plain clean dark road slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "HWY-034",
        "Highway",
        "\uace0\uc18d\ub3c4\ub85c",
        "Highway right bend plain asphalt module",
        "\uace0\uc18d\ub3c4\ub85c \uc624\ub978\ucabd \ucee4\ube0c \uc544\uc2a4\ud314\ud2b8 \ubaa8\ub4c8",
        "Right bend stage path",
        "\uc624\ub978\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean right-curving asphalt slab with subtle surface wear only",
        "Single right curve highway path module, plain clean dark road slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "HWY-035",
        "Highway",
        "\uace0\uc18d\ub3c4\ub85c",
        "Highway narrowing plain asphalt module",
        "\uace0\uc18d\ub3c4\ub85c \ud569\ub958 \uc881\uc544\uc9c0\ub294 \uae38 \ubaa8\ub4c8",
        "Narrowing stage path",
        "\uc881\uc544\uc9c0\ub294 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean highway merge slab narrowing into one runner lane, subtle asphalt wear only",
        "Single highway lane merge bottleneck module, plain clean dark asphalt body, narrowing shape, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "HWY-036",
        "Highway",
        "\uace0\uc18d\ub3c4\ub85c",
        "Highway split plain asphalt module",
        "\uace0\uc18d\ub3c4\ub85c \ubd84\uae30 \ub098\ub4e4\ubaa9 \uae38 \ubaa8\ub4c8",
        "Split stage path",
        "\ubd84\uae30 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean Y-shaped highway exit slab, subtle asphalt wear only",
        "Single split highway exit lane module, plain clean Y-shaped asphalt branch, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "RST-024",
        "Rest Stop",
        "\ud734\uac8c\uc18c",
        "Rest stop straight plain pavement module",
        "\ud734\uac8c\uc18c \uc9c1\uc120 \uc9c4\uc785\ub85c \uae38 \ubaa8\ub4c8",
        "Straight stage path",
        "\uc9c1\uc120 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean rest-stop service-road slab, muted concrete asphalt, subtle wear only",
        "Single straight rest-stop service road module, plain clean muted pavement slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "RST-025",
        "Rest Stop",
        "\ud734\uac8c\uc18c",
        "Rest stop left bend plain pavement module",
        "\ud734\uac8c\uc18c \uc67c\ucabd \ucee4\ube0c \ud3ec\uc7a5 \uae38 \ubaa8\ub4c8",
        "Left bend stage path",
        "\uc67c\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean left-curving rest-stop pavement slab, subtle wear only",
        "Single left curve rest-stop pavement module, plain clean muted slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "RST-026",
        "Rest Stop",
        "\ud734\uac8c\uc18c",
        "Rest stop right bend plain pavement module",
        "\ud734\uac8c\uc18c \uc624\ub978\ucabd \ucee4\ube0c \ud3ec\uc7a5 \uae38 \ubaa8\ub4c8",
        "Right bend stage path",
        "\uc624\ub978\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean right-curving rest-stop pavement slab, subtle wear only",
        "Single right curve rest-stop pavement module, plain clean muted slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "RST-027",
        "Rest Stop",
        "\ud734\uac8c\uc18c",
        "Rest stop narrowing plain pavement module",
        "\ud734\uac8c\uc18c \uc881\uc544\uc9c0\ub294 \ud3ec\uc7a5 \uae38 \ubaa8\ub4c8",
        "Narrowing stage path",
        "\uc881\uc544\uc9c0\ub294 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean rest-stop pavement slab narrowing into one runner lane, subtle wear only",
        "Single narrowing rest-stop pavement module, plain clean muted slab, narrowing shape, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "RST-028",
        "Rest Stop",
        "\ud734\uac8c\uc18c",
        "Rest stop split plain pavement module",
        "\ud734\uac8c\uc18c \ubd84\uae30 \ud3ec\uc7a5 \uae38 \ubaa8\ub4c8",
        "Split stage path",
        "\ubd84\uae30 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean split rest-stop pavement slab, subtle wear only",
        "Single split rest-stop pavement module, plain clean Y-shaped muted slab, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "CITY-030",
        "City",
        "\ub3c4\uc2dc",
        "City straight plain asphalt module",
        "\ub3c4\uc2dc \uc9c1\uc120 \uc544\uc2a4\ud314\ud2b8 \uae38 \ubaa8\ub4c8",
        "Straight stage path",
        "\uc9c1\uc120 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean urban asphalt slab, subtle street wear only",
        "Single straight city street module, plain clean urban asphalt body, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "CITY-031",
        "City",
        "\ub3c4\uc2dc",
        "City left bend plain asphalt module",
        "\ub3c4\uc2dc \uc67c\ucabd \ud68c\uc804 \uac70\ub9ac \ubaa8\ub4c8",
        "Left bend stage path",
        "\uc67c\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean left-turn urban asphalt slab, subtle street wear only",
        "Single left-turn city street module, plain clean urban asphalt body, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "CITY-032",
        "City",
        "\ub3c4\uc2dc",
        "City right bend plain asphalt module",
        "\ub3c4\uc2dc \uc624\ub978\ucabd \ud68c\uc804 \uac70\ub9ac \ubaa8\ub4c8",
        "Right bend stage path",
        "\uc624\ub978\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean right-turn urban asphalt slab, subtle street wear only",
        "Single right-turn city street module, plain clean urban asphalt body, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "CITY-033",
        "City",
        "\ub3c4\uc2dc",
        "City narrowing plain asphalt module",
        "\ub3c4\uc2dc \uc881\uc544\uc9c0\ub294 \uc544\uc2a4\ud314\ud2b8 \uae38 \ubaa8\ub4c8",
        "Narrowing stage path",
        "\uc881\uc544\uc9c0\ub294 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean urban asphalt slab narrowing into one runner lane, subtle street wear only",
        "Single narrowing city street module, plain clean urban asphalt body, narrowing shape, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "CITY-034",
        "City",
        "\ub3c4\uc2dc",
        "City split plain asphalt module",
        "\ub3c4\uc2dc \ubd84\uae30 \ub3c4\uc2ec \uac70\ub9ac \ubaa8\ub4c8",
        "Split stage path",
        "\ubd84\uae30 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean split downtown asphalt slab, subtle street wear only",
        "Single split downtown city street module, plain clean Y-shaped urban asphalt branch, subtle surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "GNG-031",
        "Gangnam",
        "\uac15\ub0a8",
        "Gangnam straight plain boulevard module",
        "\uac15\ub0a8 \uc9c1\uc120 \ub300\ub85c \ubaa8\ub4c8",
        "Straight stage path",
        "\uc9c1\uc120 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean glossy dark boulevard slab, subtle polished surface wear only",
        "Single straight Gangnam boulevard module, plain clean glossy dark road body, subtle polished surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "GNG-032",
        "Gangnam",
        "\uac15\ub0a8",
        "Gangnam left bend plain boulevard module",
        "\uac15\ub0a8 \uc67c\ucabd \ucee4\ube0c \ub300\ub85c \ubaa8\ub4c8",
        "Left bend stage path",
        "\uc67c\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean left-curving glossy dark boulevard slab, subtle polished surface wear only",
        "Single left curve Gangnam boulevard module, plain clean glossy dark road body, subtle polished surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "GNG-033",
        "Gangnam",
        "\uac15\ub0a8",
        "Gangnam right bend plain boulevard module",
        "\uac15\ub0a8 \uc624\ub978\ucabd \ucee4\ube0c \ub300\ub85c \ubaa8\ub4c8",
        "Right bend stage path",
        "\uc624\ub978\ucabd \ucee4\ube0c \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean right-curving glossy dark boulevard slab, subtle polished surface wear only",
        "Single right curve Gangnam boulevard module, plain clean glossy dark road body, subtle polished surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "GNG-034",
        "Gangnam",
        "\uac15\ub0a8",
        "Gangnam narrowing plain boulevard module",
        "\uac15\ub0a8 \uc881\uc544\uc9c0\ub294 \ub300\ub85c \ubaa8\ub4c8",
        "Narrowing stage path",
        "\uc881\uc544\uc9c0\ub294 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean glossy dark boulevard slab narrowing into one runner lane, subtle polished surface wear only",
        "Single narrowing Gangnam boulevard module, plain clean glossy dark road body, narrowing shape, subtle polished surface wear, no extra markings or props, white background, 3/4 view.",
    ),
    road_module_asset(
        "GNG-035",
        "Gangnam",
        "\uac15\ub0a8",
        "Gangnam split plain boulevard module",
        "\uac15\ub0a8 \ubd84\uae30 \ub300\ub85c \ubaa8\ub4c8",
        "Split stage path",
        "\ubd84\uae30 \uc2a4\ud14c\uc774\uc9c0 \uae38",
        "Clean split glossy dark boulevard slab, subtle polished surface wear only",
        "Single split Gangnam boulevard module, plain clean Y-shaped glossy dark road body, subtle polished surface wear, no extra markings or props, white background, 3/4 view.",
    ),
]


REUSE_AUDIT = [
    ("stage_02_*_highway", "버스", "Reuse", "130_HWY-017_Bus_obstacle", "고속도로와 도심 버스는 기존 버스 장애물을 색상 변형으로 재사용"),
    ("stage_02_4_highway", "요금소 본체", "Reuse", "122_HWY-009_Tollgate_booth_module; 123_HWY-010_Electronic_toll_gate", "요금소 큰 구조는 기존 항목 재사용, 차단봉/신호/결제 단말만 신규 분리"),
    ("stage_02_*_highway", "트럭/승용차/가드레일/카메라", "Reuse", "129_HWY-016; 131_HWY-018; 052_BND-004; 015_HWY-005", "핵심 교통 소품은 기존 001-181 범위에서 커버"),
    ("stage_03_*_rest_stop", "휴게소 간판/주유소/EV 충전기/벤치", "Reuse", "137_RST-008; 138_RST-009; 139_RST-010; 140_RST-011; 147_RST-018", "큰 휴게소 표식은 기존 항목 재사용"),
    ("stage_04_*_city", "버스정류장/신호등/택시/스쿠터/소화전", "Reuse", "025_CITY-001; 027_CITY-003; 161_CITY-022; 151_CITY-012; 157_CITY-018", "도심 반복 소품은 기존 항목 재사용"),
    ("stage_05_*_gangnam", "세단/쇼윈도/마네킹/백화점 입구/최종 신발", "Reuse", "172_GNG-018; 164_GNG-010; 165_GNG-011; 066_BLD-010; 176_GNG-022", "강남 핵심 분위기는 기존 항목 재사용"),
    ("stage_01_*_noryangjin", "나무 X 바리케이드", "New", "193_NRY-036_Fish-market_wooden_X_barricade", "기존 도로 공사 바리케이드와 재질/세계관이 달라 신규 분리"),
    ("stage_01_5_noryangjin", "항구 차선 신호 게이트", "New", "194_NRY-037_Harbor_lane_signal_gantry", "항구-도로 전환 구간의 방향 신호가 별도 소품으로 반복됨"),
    ("stage_02_2_highway", "이동식 LED 화살표 트레일러", "New", "195_HWY-023_Portable_LED_arrow_trailer", "기존 전광판은 고정형이라 공사 트레일러형을 신규 분리"),
    ("stage_02_2_highway", "초록 탄성 차선 분리봉", "New", "196_HWY-024_Green_flexible_lane_bollard_row", "연속 경계물로 반복 배치 가능"),
    ("stage_02_4_highway", "요금소 차단봉/차선 신호/결제 단말", "New", "197_HWY-025; 198_HWY-026; 199_HWY-027", "요금소 하위 판독 요소를 개별 생성"),
    ("stage_03_*_rest_stop", "분리수거통/카페 입간판/EV 바닥 표시", "New", "200_RST-019; 201_RST-020; 202_RST-021", "휴게소 디테일 소품을 보강"),
    ("stage_04_*_city", "보행자 신호 버튼/굴착기 팔/세로 배너", "New", "203_CITY-024; 204_CITY-025; 205_CITY-026", "도심 스테이지에 반복되는 보행/공사/거리 리듬 요소 보강"),
    ("stage_05_*_gangnam", "고급 배너/로프/주얼리/핸드백/다이아 로고/벽부등", "New", "206_GNG-023 ~ 211_GNG-028", "강남 후반부에서 보이는 프리미엄 진열과 입구 소품 보강"),
    ("stage_02_*_highway", "파란 버스/빨간 승용차/초록 승용차/노란 트럭", "New", "212_HWY-028 ~ 215_HWY-031", "기존 차량 베이스 외에 스테이지에 보이는 차종을 개별 모델로 보강"),
    ("stage_03_*_rest_stop", "주차 SUV/셔틀버스", "New", "216_RST-022; 217_RST-023", "휴게소 주차장 차량을 개별 소품으로 보강"),
    ("stage_04_*_city", "도심 버스/노란 택시/배달 라이더 스쿠터", "New", "218_CITY-027 ~ 220_CITY-029", "도심 교통 장애물 변형을 개별 모델로 보강"),
    ("stage_05_*_gangnam", "검은 고급 세단/발렛 차량", "New", "221_GNG-029; 222_GNG-030", "강남 후반부 차량을 장애물/장식 양쪽으로 사용할 수 있게 보강"),
]


REUSE_AUDIT.extend(
    [
        ("stage_01_*_noryangjin", "stage-specific road modules", "New", "NRY-038 ~ NRY-042", "Straight, left curve, right curve, narrowing, and split path modules for the Noryangjin stage."),
        ("stage_02_*_highway", "stage-specific road modules", "New", "HWY-032 ~ HWY-036", "Straight, left curve, right curve, narrowing, and split path modules for the Highway stage."),
        ("stage_03_*_rest_stop", "stage-specific road modules", "New", "RST-024 ~ RST-028", "Straight, left curve, right curve, narrowing, and split path modules for the Rest Stop stage."),
        ("stage_04_*_city", "stage-specific road modules", "New", "CITY-030 ~ CITY-034", "Straight, left curve, right curve, narrowing, and split path modules for the City stage."),
        ("stage_05_*_gangnam", "stage-specific road modules", "New", "GNG-031 ~ GNG-035", "Straight, left curve, right curve, narrowing, and split path modules for the Gangnam stage."),
    ]
)


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def asset_row(asset: dict[str, str], korean: bool) -> list[str]:
    if korean:
        return [
            asset["asset_id"],
            asset["region_kr"],
            asset["name_kr"],
            asset["name_en"],
            asset["category_kr"],
            asset["use_kr"],
            asset["priority"],
            asset["meshy_kr"],
            asset["method"],
            asset["reuse_kr"],
            asset["difficulty_kr"],
            asset["visual_notes"],
            asset["brief"],
        ]
    return [
        asset["asset_id"],
        asset["region_en"],
        asset["name_kr"],
        asset["name_en"],
        asset["category_en"],
        asset["use_en"],
        asset["priority"],
        asset["meshy_en"],
        asset["method"],
        asset["reuse_en"],
        asset["difficulty_en"],
        asset["visual_notes"],
        asset["brief"],
    ]


def upsert_assets(ws, korean: bool) -> None:
    existing = {
        str(ws.cell(row=row, column=1).value): row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=1).value
    }
    template_row = ws.max_row
    for asset in ADDITIONAL_ASSETS + ROAD_MODULE_ASSETS:
        row_values = asset_row(asset, korean)
        target_row = existing.get(asset["asset_id"])
        if target_row is None:
            target_row = ws.max_row + 1
            copy_row_style(ws, template_row, target_row)
            existing[asset["asset_id"]] = target_row
        for col, value in enumerate(row_values, start=1):
            ws.cell(target_row, col, value)


def normalize_existing_categories(ws, korean: bool) -> None:
    prop = "소품" if korean else "Prop"
    road_tile = "바닥 타일" if korean else "Road Tile"
    for row in range(2, ws.max_row + 1):
        asset_id = str(ws.cell(row=row, column=1).value)
        if asset_id in STAGE_PROP_ASSET_IDS:
            ws.cell(row=row, column=5, value=prop)
        if asset_id in STAGE_ROAD_TILE_ASSET_IDS:
            ws.cell(row=row, column=5, value=road_tile)
        if asset_id in EXISTING_ROAD_NAME_OVERRIDES:
            name_kr, name_en = EXISTING_ROAD_NAME_OVERRIDES[asset_id]
            ws.cell(row=row, column=3, value=name_kr)
            ws.cell(row=row, column=4, value=name_en)
        if asset_id in EXISTING_ROAD_NOTE_OVERRIDES:
            visual_notes, brief = EXISTING_ROAD_NOTE_OVERRIDES[asset_id]
            ws.cell(row=row, column=12, value=visual_notes)
            ws.cell(row=row, column=13, value=brief)


def existing_image_targets() -> set[str]:
    return {path.name for path in IMAGE_DIR.glob("*.png")}


def safe_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return re.sub(r"_+", "_", cleaned)


def asset_number(asset_id: str) -> str:
    return asset_id.split("-", 1)[1] if "-" in asset_id else asset_id


def asset_stage_sort_key(row: tuple[object, ...], original_index: int) -> tuple[int, int]:
    asset_id, region = row[:2]
    stage = stage_code(str(asset_id), region)
    return (STAGE_SORT_ORDER.get(stage, len(STAGE_SORT_ORDER)), original_index)


def sort_asset_rows(ws) -> None:
    indexed_rows = [
        (index, tuple(cell.value for cell in row))
        for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column))
        if row[0].value
    ]
    indexed_rows.sort(key=lambda item: asset_stage_sort_key(item[1], item[0]))

    for row_index, (_, row_values) in enumerate(indexed_rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            ws.cell(row_index, column_index, value)


def stage_code(asset_id: str, region: object) -> str:
    if asset_id in ROAD_STAGE_CODES:
        return ROAD_STAGE_CODES[asset_id]

    prefix = asset_id.split("-", 1)[0]
    if prefix in PREFIX_STAGE_CODES:
        return PREFIX_STAGE_CODES[prefix]

    region_code = STAGE_CODES.get(str(region))
    if region_code:
        return region_code

    return "COMMON"


def kind_code(asset_id: str, category: object) -> str:
    prefix = asset_id.split("-", 1)[0]
    category_text = str(category)

    if asset_id in STAGE_ROAD_TILE_ASSET_IDS:
        return "ROAD"
    if prefix == "BG" or category_text in ("Background", "배경"):
        return "BACKGROUND"
    if prefix == "ROAD" or category_text in ("Road Tile", "Road Module", "도로 모듈", "바닥 타일", "환경 모듈"):
        return "ROAD"
    if prefix == "BND" or category_text in ("Boundary", "경계물"):
        return "BOUNDARY"
    if prefix == "BLD" or category_text in ("Building Facade", "건물 파사드"):
        return "BUILDING"
    if prefix == "MOD" or category_text in ("Gameplay Module", "게임플레이 모듈", "Level Module", "레벨 모듈"):
        return "GAMEPLAY"
    if category_text in ("Obstacle", "장애물"):
        return "OBSTACLE"
    if category_text in ("Pickup", "픽업"):
        return "PICKUP"
    if category_text in ("Enemy", "적"):
        return "ENEMY"
    if category_text in ("Landmark", "랜드마크"):
        return "LANDMARK"

    return "PROPS"


def target_filename(sequence: int, row: tuple[object, ...]) -> str:
    asset_id, region, _, name_en, category = row[:5]
    return (
        f"{sequence:03d}_"
        f"{stage_code(str(asset_id), region)}_"
        f"{kind_code(str(asset_id), category)}_"
        f"{asset_number(str(asset_id))}_"
        f"{safe_name(str(name_en))}.png"
    )


def make_queue_sheet(wb, ws, korean: bool) -> None:
    title = "이미지생성대기열" if korean else "ImageGenerationQueue"
    if title in wb.sheetnames:
        del wb[title]
    queue = wb.create_sheet(title)
    headers = ["번호", "소품 코드", "영문명", "대상 파일명", "상태"] if korean else ["Sequence", "Asset ID", "Name EN", "Target Filename", "Status"]
    queue.append(headers)
    image_targets = existing_image_targets()
    for seq, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        asset_id, _, _, name_en = row[:4]
        if not asset_id:
            continue
        target = target_filename(seq, row)
        if korean:
            status = "생성됨" if target in image_targets else "이미지 없음"
        else:
            status = "Generated" if target in image_targets else "Missing image"
        queue.append([seq, asset_id, name_en, target, status])
    format_table(queue)


def make_audit_sheet(wb, korean: bool) -> None:
    title = "스테이지소품감사" if korean else "StageReferenceAudit"
    if title in wb.sheetnames:
        del wb[title]
    audit = wb.create_sheet(title)
    headers = ["스테이지", "보이는 소품", "처리", "재사용/신규 소품 ID", "근거"] if korean else ["Stage", "Visible Prop", "Action", "Reuse/New Asset ID", "Rationale"]
    audit.append(headers)
    for row in REUSE_AUDIT:
        audit.append(row)
    format_table(audit)


def format_table(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 52)
    ws.freeze_panes = "A2"


def update_current_workbook(path: Path, korean: bool) -> None:
    wb = load_workbook(path)
    ws = wb.worksheets[1]
    upsert_assets(ws, korean)
    normalize_existing_categories(ws, korean)
    sort_asset_rows(ws)
    make_audit_sheet(wb, korean)
    make_queue_sheet(wb, ws, korean)
    wb.save(path)


def write_legacy_rebuilt_sheet(path: Path, source_path: Path) -> None:
    legacy = load_workbook(path)
    source = load_workbook(source_path, read_only=True, data_only=True)
    source_ws = source.worksheets[1]

    title = "재구축목록_20260510"
    if title in legacy.sheetnames:
        del legacy[title]
    rebuilt = legacy.create_sheet(title)
    for row in source_ws.iter_rows(values_only=True):
        rebuilt.append(list(row))
    format_table(rebuilt)

    audit_title = "재사용감사_20260510"
    if audit_title in legacy.sheetnames:
        del legacy[audit_title]
    audit = legacy.create_sheet(audit_title)
    audit.append(["스테이지", "보이는 소품", "처리", "재사용/신규 소품 ID", "근거"])
    for row in REUSE_AUDIT:
        audit.append(row)
    format_table(audit)

    legacy.save(path)


def rewrite_prompt_jsonl(source_path: Path, out_path: Path) -> None:
    wb = load_workbook(source_path, read_only=True, data_only=True)
    ws = wb.worksheets[1]
    records = []
    for sequence, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue
        records.append(
            {
                "sequence": sequence,
                "asset_id": row[0],
                "region": row[1],
                "name_kr": row[2],
                "name_en": row[3],
                "type": row[4],
                "role": row[5],
                "priority": row[6],
                "meshy_fit": row[7],
                "recommended_input": row[8],
                "visual_notes": row[11],
                "brief": row[12],
            }
        )
    with out_path.open("w", encoding="utf-8", newline="\n") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def main() -> None:
    english = DESIGN_DIR / "tralalero_meshy_asset_plan.xlsx"
    korean = DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx"
    update_current_workbook(english, korean=False)
    update_current_workbook(korean, korean=True)
    rewrite_prompt_jsonl(korean, DESIGN_DIR / "meshy_image_prompts_kr.jsonl")

    for path in DESIGN_DIR.glob("*.xlsx"):
        if "MeshyAI" in path.name and not path.name.startswith("old_"):
            write_legacy_rebuilt_sheet(path, korean)


if __name__ == "__main__":
    main()
