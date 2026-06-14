from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

from rebuild_meshy_design_assets import ASSET_ID_PATTERN, DESIGN_DIR, IMAGE_DIR, target_filename


def filename_suffix_without_sequence(filename: str) -> str:
    parts = filename.split("_", 1)
    return parts[1] if len(parts) == 2 else filename


def find_current_image(sequence: int, asset_id: str, target_name: str) -> Path | None:
    target_path = IMAGE_DIR / target_name
    if target_path.exists():
        return target_path

    target_parts = target_name.split("_", 5)
    if len(target_parts) == 6:
        stage_kind_asset_prefix = "_".join(target_parts[:5]) + "_"
        for path in IMAGE_DIR.glob("*.png"):
            if path.name.startswith(stage_kind_asset_prefix):
                return path

    target_suffix = filename_suffix_without_sequence(target_name)
    for path in IMAGE_DIR.glob("*.png"):
        if filename_suffix_without_sequence(path.name) == target_suffix:
            return path

    prefix = f"{sequence:03d}_"
    candidates = [
        path
        for path in IMAGE_DIR.glob("*.png")
        if path.name.startswith(prefix) and f"_{asset_id}_" in path.name
    ]
    if candidates:
        return candidates[0]

    for path in IMAGE_DIR.glob("*.png"):
        match = ASSET_ID_PATTERN.search(path.name)
        if match and match.group(1) == asset_id:
            return path

    old_pattern = re.compile(rf"^{sequence:03d}_{re.escape(asset_id)}_.*\.png$")
    for path in IMAGE_DIR.glob("*.png"):
        if old_pattern.match(path.name):
            return path

    return None


def iter_renames() -> list[tuple[Path, Path]]:
    workbook = load_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", read_only=True, data_only=True)
    worksheet = workbook.worksheets[1]
    renames: list[tuple[Path, Path]] = []

    for sequence, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        asset_id = row[0]
        if not asset_id:
            continue

        target_name = target_filename(sequence, row)
        current = find_current_image(sequence, str(asset_id), target_name)
        if current is None:
            continue

        target = IMAGE_DIR / target_name
        if current != target:
            renames.append((current, target))

    contact_sheet = IMAGE_DIR / "001_066_contact_sheet.png"
    if contact_sheet.exists():
        renames.append((contact_sheet, IMAGE_DIR / "CONTACT_SHEET_001_066.png"))

    return renames


def main() -> None:
    parser = argparse.ArgumentParser(description="Rename Meshy design PNGs to the stage/type filename convention.")
    parser.add_argument("--apply", action="store_true", help="Apply the rename operations. Omit for preview only.")
    args = parser.parse_args()

    renames = iter_renames()
    for source, target in renames:
        print(f"{source.name} -> {target.name}")

    if not args.apply:
        print(f"Preview only: {len(renames)} rename(s). Use --apply to rename files.")
        return

    for source, target in renames:
        if target.exists():
            raise FileExistsError(f"Target already exists: {target}")
        source.rename(target)

    print(f"Applied {len(renames)} rename(s).")


if __name__ == "__main__":
    main()
