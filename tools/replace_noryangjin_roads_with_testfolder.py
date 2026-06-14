from __future__ import annotations

import json
import math
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import bpy
from mathutils import Vector
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMAGE_DIR = ROOT / "output" / "meshy_images"
DESIGN_DIR = ROOT / "docs" / "design"

MODELS_STAGE_DIR = ROOT / "Assets" / "ShooterSurvival" / "Models" / "MeshyAI" / "Stage01_Noryangjin"
TEXTURES_STAGE_DIR = ROOT / "Assets" / "ShooterSurvival" / "Textures" / "MeshyAI" / "Stage01_Noryangjin"
MATERIALS_STAGE_DIR = ROOT / "Assets" / "ShooterSurvival" / "Materials" / "MeshyAI" / "Stage01_Noryangjin"
PREFABS_STAGE_DIR = ROOT / "Assets" / "ShooterSurvival" / "Prefabs" / "MeshyAI" / "Stage01_Noryangjin"
TEST_FOLDER = ROOT / "Assets" / "ShooterSurvival" / "Models" / "MeshyAI" / "TestFolder"

NUMBERED_PNG = re.compile(r"^(?P<seq>\d{3})_(?P<rest>.+)\.png$")
NORYANGJIN_ROAD = re.compile(r"^(?:0(?:4[6-9]|5[0-4])_)?STAGE01_NRY_ROAD_0(?:3[8-9]|4[0-6])_.+")
ASSET_REF = re.compile(r"(?:\d{3}_)?([A-Z]+-\d{3})(?:_[A-Za-z0-9][A-Za-z0-9_-]*)?")


@dataclass(frozen=True)
class NewRoad:
    asset_id: str
    sequence: int
    source_fbx: str
    name_en: str
    name_kr: str
    role_en: str
    role_kr: str

    @property
    def asset_number(self) -> str:
        return self.asset_id.split("-")[1]

    @property
    def slug(self) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "_", self.name_en).strip("_")

    @property
    def base_name(self) -> str:
        return f"{self.sequence:03d}_STAGE01_NRY_ROAD_{self.asset_number}_{self.slug}"

    @property
    def image_name(self) -> str:
        return f"{self.base_name}.png"


NEW_ROADS = [
    NewRoad(
        "NRY-038",
        46,
        "Road_Straight.fbx",
        "Noryangjin modular straight timber road module",
        "노량진 모듈형 직선 목재 길",
        "Straight map-tool road module",
        "맵툴용 직선 길 모듈",
    ),
    NewRoad(
        "NRY-039",
        47,
        "Road_Left90.fbx",
        "Noryangjin modular left 90 timber road module",
        "노량진 모듈형 왼쪽 90도 목재 길",
        "Left 90-degree map-tool road module",
        "맵툴용 왼쪽 90도 길 모듈",
    ),
    NewRoad(
        "NRY-040",
        48,
        "Road_Right90.fbx",
        "Noryangjin modular right 90 timber road module",
        "노량진 모듈형 오른쪽 90도 목재 길",
        "Right 90-degree map-tool road module",
        "맵툴용 오른쪽 90도 길 모듈",
    ),
]

DROP_ASSET_IDS = {f"NRY-{number:03d}" for number in range(41, 47)}


def ensure_inside(path: Path, root: Path) -> None:
    resolved = path.resolve()
    root_resolved = root.resolve()
    if resolved != root_resolved and root_resolved not in resolved.parents:
        raise RuntimeError(f"Refusing to operate outside {root_resolved}: {resolved}")


def backup_path(path: Path, backup_root: Path) -> Path:
    return backup_root / path.name


def move_with_meta(path: Path, backup_root: Path) -> None:
    if not path.exists():
        return
    ensure_inside(path, ROOT)
    backup_root.mkdir(parents=True, exist_ok=True)
    target = backup_path(path, backup_root)
    if target.exists():
        raise FileExistsError(target)
    shutil.move(str(path), str(target))

    meta = Path(str(path) + ".meta")
    if meta.exists():
        meta_target = backup_path(meta, backup_root)
        if meta_target.exists():
            raise FileExistsError(meta_target)
        shutil.move(str(meta), str(meta_target))


def backup_existing_stage_roads(stamp: str) -> None:
    for stage_dir in [MODELS_STAGE_DIR, TEXTURES_STAGE_DIR, MATERIALS_STAGE_DIR, PREFABS_STAGE_DIR]:
        backup_root = stage_dir / "_old" / f"replaced_by_testfolder_roads_{stamp}"
        for child in sorted(stage_dir.iterdir()):
            if child.name.startswith("_"):
                continue
            if child.is_dir() and NORYANGJIN_ROAD.match(child.name):
                move_with_meta(child, backup_root)

    image_backup = IMAGE_DIR / "old" / f"noryangjin_roads_replaced_by_testfolder_{stamp}"
    for image in sorted(IMAGE_DIR.glob("*.png")):
        match = NUMBERED_PNG.match(image.name)
        if not match:
            continue
        sequence = int(match.group("seq"))
        if 46 <= sequence <= 54 and "_STAGE01_NRY_ROAD_" in image.name:
            move_with_meta(image, image_backup)


def copy_new_model_assets() -> None:
    base_texture = TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture.png"
    textures = {
        "BaseColor": base_texture,
        "Normal": TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture_normal.png",
        "Metallic": TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture_metallic.png",
        "Roughness": TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture_roughness.png",
        "Emission": TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture_emit.png",
        "Metallic_Roughness": TEST_FOLDER / "Meshy_AI_Riveted_Timber_Letter_0531105458_texture_metallic_roughness.png",
    }

    for road in NEW_ROADS:
        target_dir = MODELS_STAGE_DIR / road.base_name
        target_dir.mkdir(parents=True, exist_ok=True)

        source_fbx = TEST_FOLDER / road.source_fbx
        if not source_fbx.exists():
            raise FileNotFoundError(source_fbx)
        shutil.copy2(source_fbx, target_dir / f"{road.base_name}.fbx")

        for kind, source in textures.items():
            if source.exists():
                shutil.copy2(source, target_dir / f"{road.base_name}_{kind}.png")


def object_bounds(obj: bpy.types.Object) -> tuple[Vector, Vector]:
    points = [obj.matrix_world @ Vector(corner) for corner in obj.bound_box]
    mins = Vector((min(point[i] for point in points) for i in range(3)))
    maxs = Vector((max(point[i] for point in points) for i in range(3)))
    return mins, maxs


def look_at(obj: bpy.types.Object, target: Vector) -> None:
    direction = target - obj.location
    obj.rotation_euler = direction.to_track_quat("-Z", "Y").to_euler()


def render_preview(source_fbx: Path, target_png: Path) -> None:
    bpy.ops.wm.read_factory_settings(use_empty=True)
    try:
        bpy.ops.preferences.addon_enable(module="io_scene_fbx")
    except Exception:
        pass

    bpy.ops.import_scene.fbx(filepath=str(source_fbx))
    objects = [obj for obj in bpy.context.scene.objects if obj.type == "MESH"]
    if not objects:
        raise RuntimeError(f"No mesh objects imported from {source_fbx}")

    bpy.ops.object.select_all(action="DESELECT")
    for obj in objects:
        obj.select_set(True)
    bpy.context.view_layer.objects.active = objects[0]
    if len(objects) > 1:
        bpy.ops.object.join()
    obj = bpy.context.view_layer.objects.active

    mins, maxs = object_bounds(obj)
    center = (mins + maxs) * 0.5
    dims = maxs - mins

    light_data = bpy.data.lights.new("Key_Light", type="AREA")
    light_obj = bpy.data.objects.new("Key_Light", light_data)
    bpy.context.collection.objects.link(light_obj)
    light_obj.location = (center.x, center.y - 1.8, center.z + 1.4)
    light_data.energy = 450
    light_data.size = 3.0

    camera_data = bpy.data.cameras.new("Preview_Camera")
    camera = bpy.data.objects.new("Preview_Camera", camera_data)
    bpy.context.collection.objects.link(camera)
    camera.location = (center.x, center.y - 2.0, center.z + 0.45)
    look_at(camera, center)
    camera_data.type = "ORTHO"
    camera_data.ortho_scale = max(dims.x, dims.z) * 1.25
    bpy.context.scene.camera = camera

    world = bpy.context.scene.world or bpy.data.worlds.new("World")
    bpy.context.scene.world = world
    world.color = (0.025, 0.025, 0.025)

    available_engines = [item.identifier for item in bpy.types.RenderSettings.bl_rna.properties["engine"].enum_items]
    bpy.context.scene.render.engine = "BLENDER_EEVEE_NEXT" if "BLENDER_EEVEE_NEXT" in available_engines else "BLENDER_EEVEE"
    if hasattr(bpy.context.scene, "eevee"):
        bpy.context.scene.eevee.taa_render_samples = 32
    bpy.context.scene.render.resolution_x = 1024
    bpy.context.scene.render.resolution_y = 1024
    bpy.context.scene.render.film_transparent = False
    bpy.context.scene.view_settings.view_transform = "Filmic"
    bpy.context.scene.view_settings.look = "Medium High Contrast"
    bpy.context.scene.render.filepath = str(target_png)
    bpy.ops.render.render(write_still=True)


def render_new_previews() -> None:
    for road in NEW_ROADS:
        render_preview(TEST_FOLDER / road.source_fbx, IMAGE_DIR / road.image_name)


def renumber_following_images() -> None:
    numbered = []
    for path in IMAGE_DIR.glob("*.png"):
        match = NUMBERED_PNG.match(path.name)
        if match:
            numbered.append((int(match.group("seq")), path, match.group("rest")))

    for sequence, path, rest in sorted(numbered):
        if sequence < 55:
            continue
        target = IMAGE_DIR / f"{sequence - 6:03d}_{rest}.png"
        if target.exists():
            raise FileExistsError(f"Cannot renumber {path.name}; target exists: {target.name}")
        path.rename(target)


def active_image_names() -> list[str]:
    names = [
        path.name
        for path in IMAGE_DIR.glob("*.png")
        if path.is_file() and NUMBERED_PNG.match(path.name)
    ]
    return sorted(names, key=lambda name: int(name[:3]))


def rest_name(filename: str) -> str:
    return filename[4:]


def copy_sheet_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def rewrite_rows(ws, rows: list[list[object]], template_row: int = 2) -> None:
    max_col = ws.max_column
    template_cells = [ws.cell(row=template_row, column=column) for column in range(1, max_col + 1)]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_index, values in enumerate(rows, start=2):
        for column in range(1, max_col + 1):
            cell = ws.cell(
                row=row_index,
                column=column,
                value=values[column - 1] if column <= len(values) else None,
            )
            copy_sheet_style(template_cells[column - 1], cell)


def new_road_row(existing: list[object], road: NewRoad, korean: bool) -> list[object]:
    row = list(existing)
    row[0] = road.asset_id
    row[2] = road.name_kr
    row[3] = road.name_en
    row[5] = road.role_kr if korean else road.role_en
    row[8] = "Imported FBX + shared material"
    row[9] = "맵툴용 3개 도로 모듈 세트" if korean else "Three-piece map-tool road module set"
    row[11] = (
        "TestFolder에서 분리한 MeshyAI 목재 길 모듈. 직선/좌90/우90 세 조각만 유지하고, "
        "연결되는 끝에는 금속 캡이 없으며 공통 텍스처를 사용한다."
        if korean
        else "MeshyAI timber road module split from TestFolder. Keeps only straight, left 90, and right 90 pieces; connector ends have no metal caps and all pieces share one texture set."
    )
    row[12] = (
        f"Single {road.name_en}, imported FBX map-tool road module, wet riveted timber surface, shared texture set, open connector ends, clean background preview."
    )
    return row


def stage_prefix_to_asset_prefix(rest: str) -> str | None:
    mapping = {
        "STAGE01_NRY": "NRY",
        "STAGE02_HWY": "HWY",
        "STAGE03_RST": "RST",
        "STAGE04_CITY": "CITY",
        "STAGE05_GNG": "GNG",
        "COMMON": "COM",
    }
    for stage, prefix in mapping.items():
        if rest.startswith(stage + "_"):
            return prefix
    return None


def infer_asset_id_from_rest(rest: str) -> str | None:
    parts = rest.split("_")
    if len(parts) < 4:
        return None

    stage_code = "_".join(parts[:2]) if parts[0] != "COMMON" else "COMMON"
    if parts[0] == "STAGE01":
        stage_code = "STAGE01_NRY"
    elif parts[0] == "STAGE02":
        stage_code = "STAGE02_HWY"
    elif parts[0] == "STAGE03":
        stage_code = "STAGE03_RST"
    elif parts[0] == "STAGE04":
        stage_code = "STAGE04_CITY"
    elif parts[0] == "STAGE05":
        stage_code = "STAGE05_GNG"

    prefix = stage_prefix_to_asset_prefix(stage_code)
    if prefix is None:
        return None

    asset_number = parts[3] if parts[0] != "COMMON" else parts[2]
    if not asset_number.isdigit():
        return None
    return f"{prefix}-{int(asset_number):03d}"


def sync_audit_sheet(wb, sequence_by_asset: dict[str, int]) -> None:
    if len(wb.worksheets) < 4:
        return
    audit_ws = wb.worksheets[-2]
    for row_index in range(2, audit_ws.max_row + 1):
        cell = audit_ws.cell(row=row_index, column=4)
        if not isinstance(cell.value, str):
            continue
        cell.value = ASSET_REF.sub(
            lambda match: f"{sequence_by_asset[match.group(1)]:03d}_{match.group(1)}"
            if match.group(1) in sequence_by_asset
            else match.group(1),
            cell.value,
        )


def sync_workbook(path: Path, image_names: list[str], korean: bool) -> tuple[int, dict[str, int]]:
    wb = load_workbook(path)
    asset_ws = wb.worksheets[3] if korean else wb.worksheets[1]
    queue_ws = wb.worksheets[-1]

    rows_by_asset: dict[str, list[object]] = {}
    ordered_rows: list[list[object]] = []
    for row in asset_ws.iter_rows(min_row=2, max_col=asset_ws.max_column, values_only=True):
        if not row or not row[0]:
            continue
        values = list(row)
        asset_id = str(values[0])
        if asset_id in DROP_ASSET_IDS:
            continue
        matching_road = next((road for road in NEW_ROADS if road.asset_id == asset_id), None)
        if matching_road:
            values = new_road_row(values, matching_road, korean)
        rows_by_asset[asset_id] = values
        ordered_rows.append(values)

    queue_by_rest: dict[str, str] = {
        rest_name(str(row[3])): str(row[1])
        for row in queue_ws.iter_rows(min_row=2, values_only=True)
        if row and row[1] and row[3]
    }
    for road in NEW_ROADS:
        queue_by_rest[rest_name(road.image_name)] = road.asset_id

    queue_rows: list[list[object]] = []
    sequence_by_asset: dict[str, int] = {}
    status = "생성됨" if korean else "Generated"

    for sequence, filename in enumerate(image_names, start=1):
        rest = rest_name(filename)
        asset_id = queue_by_rest.get(rest) or infer_asset_id_from_rest(rest)
        if not asset_id or asset_id not in rows_by_asset:
            raise RuntimeError(f"{path.name}: cannot match image to asset row: {filename} -> {asset_id}")
        row = rows_by_asset[asset_id]
        queue_rows.append([sequence, asset_id, row[3], filename, status])
        sequence_by_asset[asset_id] = sequence

    rewrite_rows(asset_ws, ordered_rows)
    rewrite_rows(queue_ws, queue_rows)
    sync_audit_sheet(wb, sequence_by_asset)
    wb.save(path)
    return len(ordered_rows), sequence_by_asset


def write_korean_prompts(image_names: list[str]) -> int:
    workbook = load_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", read_only=True, data_only=True)
    asset_ws = workbook.worksheets[3]
    queue_ws = workbook.worksheets[-1]
    row_by_asset = {
        str(row[0]): row
        for row in asset_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    }
    asset_by_sequence = {
        int(row[0]): str(row[1])
        for row in queue_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0] and row[1]
    }
    output = DESIGN_DIR / "meshy_image_prompts_kr.jsonl"
    with output.open("w", encoding="utf-8", newline="\n") as handle:
        for sequence in range(1, len(image_names) + 1):
            asset_id = asset_by_sequence[sequence]
            row = row_by_asset[asset_id]
            payload = {
                "sequence": sequence,
                "asset_id": row[0],
                "region": row[1],
                "name_kr": row[2],
                "name_en": row[3],
                "type": row[4],
                "role": row[5],
                "priority": row[6],
                "meshy_fit": row[7],
                "recommended_input": row[8],
                "visual_notes": row[11],
                "brief": row[12],
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
    return len(image_names)


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_existing_stage_roads(stamp)
    copy_new_model_assets()
    render_new_previews()
    renumber_following_images()

    image_names = active_image_names()
    workbook_results = []
    workbook_results.append(
        ("tralalero_meshy_asset_plan_kr.xlsx", *sync_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", image_names, True))
    )
    old_english = DESIGN_DIR / "old" / "tralalero_meshy_asset_plan.xlsx"
    if old_english.exists():
        workbook_results.append(
            ("old/tralalero_meshy_asset_plan.xlsx", *sync_workbook(old_english, image_names, False))
        )
    prompt_count = write_korean_prompts(image_names)

    print(f"Backup stamp: {stamp}")
    print(f"Active numbered PNGs: {len(image_names)}")
    for name, rows, sequence_by_asset in workbook_results:
        print(f"{name}: asset rows={rows}, queue rows={len(sequence_by_asset)}")
    print(f"meshy_image_prompts_kr.jsonl: {prompt_count} rows")
    for road in NEW_ROADS:
        print(f"Promoted {road.asset_id}: {road.base_name}")


if __name__ == "__main__":
    main()
