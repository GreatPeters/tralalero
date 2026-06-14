from __future__ import annotations

import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DESIGN_DIR = ROOT / "docs" / "design"
IMAGE_DIR = ROOT / "output" / "meshy_images"

WORKBOOKS = [
    DESIGN_DIR / "tralalero_meshy_asset_plan.xlsx",
    DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx",
]

NUMBERED_IMAGE = re.compile(r"^\d{3}_.+\.png$")
AUDIT_ASSET_REF = re.compile(r"(?:\d{3}_)?([A-Z]+-\d{3})(?:_[A-Za-z0-9][A-Za-z0-9_-]*)?")


def active_image_names() -> list[str]:
    names = [
        path.name
        for path in IMAGE_DIR.glob("*.png")
        if path.is_file() and NUMBERED_IMAGE.match(path.name)
    ]
    return sorted(names, key=lambda name: int(name[:3]))


def rest_name(filename: str) -> str:
    return filename[4:]


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def clear_and_write_rows(ws, rows: list[list[object]], template_row: int = 2) -> None:
    max_col = ws.max_column
    template_styles = [copy(ws.cell(row=template_row, column=col)._style) for col in range(1, max_col + 1)]
    template_alignments = [copy(ws.cell(row=template_row, column=col).alignment) for col in range(1, max_col + 1)]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_index, values in enumerate(rows, start=2):
        for col_index in range(1, max_col + 1):
            cell = ws.cell(row=row_index, column=col_index, value=values[col_index - 1] if col_index <= len(values) else None)
            cell._style = copy(template_styles[col_index - 1])
            cell.alignment = copy(template_alignments[col_index - 1])


def format_audit_reference(match: re.Match[str], sequence_by_asset: dict[str, int]) -> str:
    asset_id = match.group(1)
    sequence = sequence_by_asset.get(asset_id)
    if sequence is None:
        return asset_id
    return f"{sequence:03d}_{asset_id}"


def sync_audit_sheet(wb, sequence_by_asset: dict[str, int]) -> None:
    if len(wb.worksheets) < 4:
        return

    audit_ws = wb.worksheets[3]
    if audit_ws.max_column < 4:
        return

    for row_index in range(2, audit_ws.max_row + 1):
        cell = audit_ws.cell(row=row_index, column=4)
        if not isinstance(cell.value, str):
            continue
        cell.value = AUDIT_ASSET_REF.sub(
            lambda match: format_audit_reference(match, sequence_by_asset),
            cell.value,
        )


def sync_workbook(path: Path, image_names: list[str]) -> tuple[int, dict[str, int]]:
    wb = load_workbook(path)
    asset_ws = wb.worksheets[1]
    queue_ws = wb.worksheets[-1]

    queue_by_rest: dict[str, str] = {}
    for row in queue_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or not row[3]:
            continue
        queue_by_rest[rest_name(str(row[3]))] = str(row[1])

    row_by_asset: dict[str, list[object]] = {}
    for row in asset_ws.iter_rows(min_row=2, max_col=asset_ws.max_column, values_only=True):
        if row and row[0]:
            row_by_asset[str(row[0])] = list(row)

    missing: list[str] = []
    new_asset_rows: list[list[object]] = []
    new_queue_rows: list[list[object]] = []
    new_sequence_by_asset: dict[str, int] = {}

    status_value = "생성됨" if path.name.endswith("_kr.xlsx") else "Generated"

    for sequence, filename in enumerate(image_names, start=1):
        asset_id = queue_by_rest.get(rest_name(filename))
        if not asset_id:
            missing.append(filename)
            continue
        asset_row = row_by_asset.get(asset_id)
        if not asset_row:
            missing.append(f"{filename} -> {asset_id}")
            continue

        new_asset_rows.append(asset_row)
        new_queue_rows.append([sequence, asset_id, asset_row[3], filename, status_value])
        new_sequence_by_asset[asset_id] = sequence

    if missing:
        raise RuntimeError(f"{path.name}: could not match {len(missing)} active image(s): {missing[:8]}")

    clear_and_write_rows(asset_ws, new_asset_rows)
    clear_and_write_rows(queue_ws, new_queue_rows)
    sync_audit_sheet(wb, new_sequence_by_asset)
    wb.save(path)
    return len(new_asset_rows), new_sequence_by_asset


def write_korean_prompts(image_names: list[str]) -> int:
    workbook = load_workbook(DESIGN_DIR / "tralalero_meshy_asset_plan_kr.xlsx", read_only=True, data_only=True)
    asset_ws = workbook.worksheets[1]
    queue_ws = workbook.worksheets[-1]

    row_by_asset = {
        str(row[0]): row
        for row in asset_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    }
    asset_by_sequence = {
        int(row[0]): str(row[1])
        for row in queue_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0] and row[1]
    }

    output = DESIGN_DIR / "meshy_image_prompts_kr.jsonl"
    with output.open("w", encoding="utf-8", newline="\n") as handle:
        for sequence in range(1, len(image_names) + 1):
            asset_id = asset_by_sequence[sequence]
            row = row_by_asset[asset_id]
            payload = {
                "sequence": sequence,
                "asset_id": row[0],
                "region": row[1],
                "name_kr": row[2],
                "name_en": row[3],
                "type": row[4],
                "role": row[5],
                "priority": row[6],
                "meshy_fit": row[7],
                "recommended_input": row[8],
                "visual_notes": row[11],
                "brief": row[12],
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
    return len(image_names)


def main() -> None:
    image_names = active_image_names()
    if not image_names:
        raise RuntimeError("No active numbered image PNGs found.")

    results = []
    for workbook in WORKBOOKS:
        count, _ = sync_workbook(workbook, image_names)
        results.append((workbook.name, count))
    prompt_count = write_korean_prompts(image_names)

    for workbook, count in results:
        print(f"{workbook}: {count} active rows")
    print(f"meshy_image_prompts_kr.jsonl: {prompt_count} active rows")


if __name__ == "__main__":
    main()
