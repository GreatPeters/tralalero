# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DESIGN_DIR = ROOT / "docs" / "design"

ROAD_VARIANTS_EN = {
    "straight": ("straight modular prefab piece", "straight snap-together road piece"),
    "left": ("90-degree left corner modular prefab piece", "left 90-degree snap-together corner piece"),
    "right": ("90-degree right corner modular prefab piece", "right 90-degree snap-together corner piece"),
    "narrow": ("narrowing modular connector piece", "narrowing snap-together connector piece"),
    "split": ("T-junction / split modular prefab piece", "T-junction split snap-together road piece"),
}

ROAD_VARIANTS_KR = {
    "straight": ("직선 모듈 프리팹 조각", "스냅 연결용 직선 길 조각"),
    "left": ("90도 왼쪽 코너 모듈 프리팹 조각", "스냅 연결용 왼쪽 90도 코너 조각"),
    "right": ("90도 오른쪽 코너 모듈 프리팹 조각", "스냅 연결용 오른쪽 90도 코너 조각"),
    "narrow": ("좁아지는 연결 모듈 조각", "스냅 연결용 좁아지는 길 조각"),
    "split": ("T자 / 분기 모듈 프리팹 조각", "스냅 연결용 T자 분기 길 조각"),
}

STAGE_ROWS = {
    "NRY-038": ("Noryangjin", "straight"),
    "NRY-039": ("Noryangjin", "left"),
    "NRY-040": ("Noryangjin", "right"),
    "NRY-041": ("Noryangjin", "narrow"),
    "NRY-042": ("Noryangjin", "split"),
    "HWY-032": ("Highway", "straight"),
    "HWY-033": ("Highway", "left"),
    "HWY-034": ("Highway", "right"),
    "HWY-035": ("Highway", "narrow"),
    "HWY-036": ("Highway", "split"),
    "RST-024": ("Rest stop", "straight"),
    "RST-025": ("Rest stop", "left"),
    "RST-026": ("Rest stop", "right"),
    "RST-027": ("Rest stop", "narrow"),
    "RST-028": ("Rest stop", "split"),
    "CITY-030": ("City", "straight"),
    "CITY-031": ("City", "left"),
    "CITY-032": ("City", "right"),
    "CITY-033": ("City", "narrow"),
    "CITY-034": ("City", "split"),
    "GNG-031": ("Gangnam", "straight"),
    "GNG-032": ("Gangnam", "left"),
    "GNG-033": ("Gangnam", "right"),
    "GNG-034": ("Gangnam", "narrow"),
    "GNG-035": ("Gangnam", "split"),
}

STAGE_SURFACE_EN = {
    "Noryangjin": "wet wooden pier deck with flat snap ends, raised side beams, posts, ropes, rivet strips, blue puddles, and no checkerboard floor",
    "Highway": "modular elevated asphalt road with flat snap ends, guardrails, curbs, lane markings, and clean 3D slab thickness",
    "Rest stop": "modular service-pavement road with flat snap ends, soft concrete curbs, subtle pavement seams, and clean 3D slab thickness",
    "City": "modular city asphalt road with flat snap ends, curbs, crosswalk markings, lane markings, and clean 3D slab thickness",
    "Gangnam": "modular dark boulevard road with flat snap ends, glossy asphalt, premium curb trim, lane markings, and clean 3D slab thickness",
}

STAGE_SURFACE_KR = {
    "Noryangjin": "평평한 스냅 연결 끝단, 입체 측면 보, 말뚝, 로프, 리벳 스트립, 푸른 물웅덩이가 있는 젖은 목재 부두 데크. 바둑판 바닥 없음",
    "Highway": "평평한 스냅 연결 끝단, 가드레일, 커브, 차선 표시, 입체 두께가 있는 모듈형 고속도로 아스팔트",
    "Rest stop": "평평한 스냅 연결 끝단, 부드러운 콘크리트 커브, 은은한 포장 이음선, 입체 두께가 있는 모듈형 휴게소 서비스 포장길",
    "City": "평평한 스냅 연결 끝단, 커브, 횡단보도 표시, 차선 표시, 입체 두께가 있는 모듈형 도시 아스팔트",
    "Gangnam": "평평한 스냅 연결 끝단, 광택 있는 어두운 아스팔트, 고급 커브 장식, 차선 표시, 입체 두께가 있는 모듈형 강남 대로",
}

COMMON_ROWS_EN = {
    "ROAD-007": (
        "Common modular asphalt jump-ramp piece with flat snap ends, thick wedge volume, yellow chevrons, lane markings, and 3D slab thickness.",
        "Single common modular asphalt jump-ramp prefab piece, flat snap ends, thick wedge volume, yellow chevron route cue, lane markings, clean white background, 3/4 top view.",
    ),
    "ROAD-008": (
        "Common modular asphalt bridge/transition piece with flat snap ends, multi-lane markings, side rails, and 3D slab thickness.",
        "Single common modular asphalt bridge transition prefab piece, flat snap ends, multi-lane markings, side rails, clean white background, 3/4 top view.",
    ),
    "MOD-001": (
        "Common straight modular asphalt lane piece with flat snap ends, dashed lane markings, side curbs, and 3D slab thickness.",
        "Single common straight modular asphalt road prefab piece, flat snap ends, dashed lane markings, side curbs, clean white background, 3/4 top view.",
    ),
    "MOD-002": (
        "Common 90-degree left corner modular asphalt lane piece with flat snap ends, dashed lane markings, and side curbs.",
        "Single common left 90-degree modular asphalt road prefab piece, flat snap ends, dashed lane markings, side curbs, clean white background, 3/4 top view.",
    ),
    "MOD-003": (
        "Common 90-degree right corner modular asphalt lane piece with flat snap ends, dashed lane markings, and side curbs.",
        "Single common right 90-degree modular asphalt road prefab piece, flat snap ends, dashed lane markings, side curbs, clean white background, 3/4 top view.",
    ),
    "MOD-004": (
        "Common narrowing modular asphalt connector with flat snap ends, readable bottleneck, lane markings, and side curbs.",
        "Single common narrowing modular asphalt connector prefab piece, flat snap ends, readable bottleneck, lane markings, side curbs, clean white background, 3/4 top view.",
    ),
    "MOD-005": (
        "Common obstacle-layout modular asphalt piece with flat snap ends, cones, small block obstacles, lane markings, and 3D slab thickness.",
        "Single common obstacle-layout modular asphalt prefab piece, flat snap ends, lane markings, traffic cones and block obstacles, clean white background, 3/4 top view.",
    ),
    "MOD-006": (
        "Common side-background modular dressing piece with dark center service panel, blue side blocks, and flat rectangular prefab read.",
        "Single common side-background modular dressing prefab piece, dark center panel, blue side inset blocks, clean white background, 3/4 top view.",
    ),
    "GAME-011": (
        "Common modular asphalt jump coin-line piece with flat snap ends, staggered floating coins, lane markings, and 3D slab thickness.",
        "Single common jump coin-line modular asphalt prefab piece, flat snap ends, staggered gold coins, lane markings, clean white background, 3/4 top view.",
    ),
    "GAME-012": (
        "Common modular asphalt swerve coin-line piece with flat snap ends, diagonal coin path, lane markings, and 3D slab thickness.",
        "Single common swerve coin-line modular asphalt prefab piece, flat snap ends, diagonal gold coin path, lane markings, clean white background, 3/4 top view.",
    ),
    "MOD-008": (
        "Common T-junction / split modular asphalt piece with flat snap ends, lane markings, side curbs, and clear branch choice.",
        "Single common T-junction split modular asphalt prefab piece, flat snap ends, clear branch choice, lane markings, side curbs, clean white background, 3/4 top view.",
    ),
    "MOD-009": (
        "Common sloped modular asphalt lane piece with flat snap ends, tilted slab read, lane markings, and side curbs.",
        "Single common sloped modular asphalt road prefab piece, flat snap ends, tilted slab read, lane markings, clean white background, 3/4 top view.",
    ),
    "MOD-010": (
        "Common underpass modular road piece with flat snap ends, tunnel header, hazard stripe, readable clearance, and asphalt lane through.",
        "Single common underpass modular road prefab piece, flat snap ends, dark tunnel header, hazard stripe, asphalt lane through, clean white background, 3/4 top view.",
    ),
}

COMMON_ROWS_KR = {
    "ROAD-007": (
        "평평한 스냅 연결 끝단, 두꺼운 경사 볼륨, 노란 화살표, 차선 표시, 입체 두께가 있는 공통 아스팔트 점프 램프 모듈.",
        COMMON_ROWS_EN["ROAD-007"][1],
    ),
    "ROAD-008": (
        "평평한 스냅 연결 끝단, 다차선 표시, 측면 레일, 입체 두께가 있는 공통 아스팔트 브리지/전환 모듈.",
        COMMON_ROWS_EN["ROAD-008"][1],
    ),
    "MOD-001": (
        "평평한 스냅 연결 끝단, 점선 차선, 측면 커브, 입체 두께가 있는 공통 직선 아스팔트 차선 모듈.",
        COMMON_ROWS_EN["MOD-001"][1],
    ),
    "MOD-002": (
        "평평한 스냅 연결 끝단, 점선 차선, 측면 커브가 있는 공통 90도 왼쪽 코너 아스팔트 모듈.",
        COMMON_ROWS_EN["MOD-002"][1],
    ),
    "MOD-003": (
        "평평한 스냅 연결 끝단, 점선 차선, 측면 커브가 있는 공통 90도 오른쪽 코너 아스팔트 모듈.",
        COMMON_ROWS_EN["MOD-003"][1],
    ),
    "MOD-004": (
        "평평한 스냅 연결 끝단, 읽기 쉬운 병목, 차선 표시, 측면 커브가 있는 공통 좁아지는 아스팔트 연결 모듈.",
        COMMON_ROWS_EN["MOD-004"][1],
    ),
    "MOD-005": (
        "평평한 스냅 연결 끝단, 콘, 작은 블록 장애물, 차선 표시, 입체 두께가 있는 공통 장애물 배치 아스팔트 모듈.",
        COMMON_ROWS_EN["MOD-005"][1],
    ),
    "MOD-006": (
        "어두운 중앙 서비스 패널, 파란 측면 블록, 평평한 직사각형 프리팹 형태의 공통 측면 배경 장식 모듈.",
        COMMON_ROWS_EN["MOD-006"][1],
    ),
    "GAME-011": (
        "평평한 스냅 연결 끝단, 엇갈린 금화, 차선 표시, 입체 두께가 있는 공통 아스팔트 점프 코인 라인 모듈.",
        COMMON_ROWS_EN["GAME-011"][1],
    ),
    "GAME-012": (
        "평평한 스냅 연결 끝단, 대각선 금화 경로, 차선 표시, 입체 두께가 있는 공통 아스팔트 회피 코인 라인 모듈.",
        COMMON_ROWS_EN["GAME-012"][1],
    ),
    "MOD-008": (
        "평평한 스냅 연결 끝단, 차선 표시, 측면 커브, 명확한 분기 선택이 있는 공통 T자/분기 아스팔트 모듈.",
        COMMON_ROWS_EN["MOD-008"][1],
    ),
    "MOD-009": (
        "평평한 스냅 연결 끝단, 기울어진 슬랩 느낌, 차선 표시, 측면 커브가 있는 공통 경사 아스팔트 차선 모듈.",
        COMMON_ROWS_EN["MOD-009"][1],
    ),
    "MOD-010": (
        "평평한 스냅 연결 끝단, 터널 헤더, 위험 표시 줄무늬, 명확한 통과 공간, 관통 아스팔트 차선이 있는 공통 언더패스 모듈.",
        COMMON_ROWS_EN["MOD-010"][1],
    ),
}


def stage_update_en(stage: str, variant: str) -> tuple[str, str]:
    variant_note, brief_piece = ROAD_VARIANTS_EN[variant]
    surface = STAGE_SURFACE_EN[stage]
    return (
        f"{stage} {variant_note}: {surface}.",
        f"Single {stage} {brief_piece}, modular road prefab kit style, {surface}, clean white background, 3/4 top view.",
    )


def stage_update_kr(stage: str, variant: str) -> tuple[str, str]:
    variant_note, _ = ROAD_VARIANTS_KR[variant]
    surface = STAGE_SURFACE_KR[stage]
    return (
        f"{stage} {variant_note}: {surface}.",
        stage_update_en(stage, variant)[1],
    )


def workbook_updates(korean: bool) -> dict[str, tuple[str, str]]:
    updates: dict[str, tuple[str, str]] = {}
    for asset_id, (stage, variant) in STAGE_ROWS.items():
        updates[asset_id] = stage_update_kr(stage, variant) if korean else stage_update_en(stage, variant)
    updates.update(COMMON_ROWS_KR if korean else COMMON_ROWS_EN)
    return updates


def update_workbook(filename: str, updates: dict[str, tuple[str, str]]) -> int:
    path = DESIGN_DIR / filename
    workbook = load_workbook(path)
    asset_sheet = workbook.worksheets[1]
    changed = 0
    for row in range(2, asset_sheet.max_row + 1):
        asset_id = asset_sheet.cell(row=row, column=1).value
        if asset_id not in updates:
            continue
        visual_note, brief = updates[asset_id]
        asset_sheet.cell(row=row, column=12).value = visual_note
        asset_sheet.cell(row=row, column=13).value = brief
        changed += 1
    workbook.save(path)
    return changed


def main() -> None:
    results = [
        ("tralalero_meshy_asset_plan.xlsx", update_workbook("tralalero_meshy_asset_plan.xlsx", workbook_updates(False))),
        ("tralalero_meshy_asset_plan_kr.xlsx", update_workbook("tralalero_meshy_asset_plan_kr.xlsx", workbook_updates(True))),
    ]
    for workbook, count in results:
        print(f"{workbook}: updated {count} modular road note row(s)")


if __name__ == "__main__":
    main()
